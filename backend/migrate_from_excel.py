"""
VAIRAM FINCORP — Excel → PostgreSQL Migration Script
=====================================================
Wipes all PostgreSQL tables and reloads cleanly from the latest Diamond Fincorp master workbook.

Usage:
    python migrate_from_excel.py                    # uses default paths
    python migrate_from_excel.py --source <path>    # override source Excel file
    python migrate_from_excel.py --dry-run          # validate without writing

Requires:  DATABASE_URL environment variable set to the PostgreSQL connection string.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from sqlalchemy import create_engine, text

# ---------------------------------------------------------------------------
# Defaults
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
SOURCE_DEFAULT = PROJECT_ROOT / "Diamond Fincorp Master Data.xlsx"
LEGACY_SOURCE_DEFAULT = PROJECT_ROOT / "excel_schema" / "DIAMOND FINANCE DATA.xlsm"
if not SOURCE_DEFAULT.exists():
    SOURCE_DEFAULT = LEGACY_SOURCE_DEFAULT

VALID_LOAN_STATUSES = {"ACTIVE", "COMPLETED", "HELP", "LOSS", "CLOSED"}
VALID_PAYMENT_TYPES = {"INTEREST", "PRINCIPAL", "BOTH", "BALANCE"}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text_val = str(value).strip().replace(",", "")
    if not text_val:
        return default
    try:
        return float(text_val)
    except ValueError:
        return default


def to_datetime(value: Any) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw[:19], fmt)
        except ValueError:
            continue
    return None


def normalize_phone(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def iter_nonempty_rows(ws) -> Tuple[List[Any], List[Tuple[int, tuple]]]:
    """Return (headers, [(row_number, row_values), ...]) skipping blank-key rows."""
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and row[0] not in (None, ""):
            rows.append((row_number, row))
    return headers, rows


# ---------------------------------------------------------------------------
# Extract records from the source workbook
# ---------------------------------------------------------------------------

def load_source_records(source_file: Path) -> Dict[str, List[Dict[str, Any]]]:
    """Read the source workbook and return normalised record dicts."""
    print(f"  Opening workbook: {source_file}")
    wb = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    print(f"  Available sheets: {wb.sheetnames}")

    # Sheet names have trailing spaces in the actual workbook
    borrower_sheet = None
    loan_sheet = None
    payment_sheet = "Payment_Transactions"

    for name in wb.sheetnames:
        if name.strip() == "Borrower_Master":
            borrower_sheet = name
        if name.strip() == "Loan_Master":
            loan_sheet = name

    if not borrower_sheet:
        raise ValueError("Borrower_Master sheet not found in workbook")
    if not loan_sheet:
        raise ValueError("Loan_Master sheet not found in workbook")
    if payment_sheet not in wb.sheetnames:
        raise ValueError("Payment_Transactions sheet not found in workbook")

    # ---- Customers ----
    _, customer_rows = iter_nonempty_rows(wb[borrower_sheet])
    customers: List[Dict[str, Any]] = []
    skipped_customers = 0
    for row_number, row in customer_rows:
        customer_id = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        if not customer_id or not name:
            skipped_customers += 1
            continue
        is_active = str(row[4] or "").strip().lower()
        status = "ACTIVE" if is_active in {"yes", "true", "1"} else "INACTIVE"
        customers.append({
            "customer_id": customer_id,
            "name": name,
            "phone": normalize_phone(row[2]),
            "email": "",
            "address": str(row[3] or "").strip(),
            "id_proof_type": "",
            "id_proof_number": "",
            "status": status,
            "created_date": to_datetime(row[5]) or datetime.now(),
            "notes": "",
        })
    if skipped_customers:
        print(f"  ⚠ Skipped {skipped_customers} customer rows with missing ID/name")

    # ---- Loans ----
    _, loan_rows = iter_nonempty_rows(wb[loan_sheet])
    loans: List[Dict[str, Any]] = []
    skipped_loans = 0
    for row_number, row in loan_rows:
        loan_id = str(row[0] or "").strip()
        customer_id = str(row[1] or "").strip()
        if not loan_id or not customer_id:
            skipped_loans += 1
            continue
        raw_status = str(row[8] or "ACTIVE").strip().upper() or "ACTIVE"
        if raw_status not in VALID_LOAN_STATUSES:
            print(f"  ⚠ Loan {loan_id}: Unknown status '{raw_status}' → recovered to ACTIVE")
            raw_status = "ACTIVE"
        add_on_principal = round(max(0.0, to_float(row[10])), 2)
        principal = round(to_float(row[4]), 2)
        fresh_principal = round(max(0.0, principal - add_on_principal), 2)
        loans.append({
            "loan_id": loan_id,
            "customer_id": customer_id,
            "principal_amount": principal,
            "add_on_principal": add_on_principal,
            "fresh_principal": fresh_principal,
            "interest_rate": round(to_float(row[5]), 6),
            "loan_type": "PERSONAL",
            "transaction_type": str(row[3] or "DEBT").strip().upper() or "DEBT",
            "start_date": to_datetime(row[6]),
            "tenure_months": None,
            "status": raw_status,
            "fund_source": str(row[7] or "").strip(),
            "created_date": to_datetime(row[9]) or to_datetime(row[6]) or datetime.now(),
            "closed_date": None,
            "notes": str(row[11] or "").strip() if len(row) > 11 else "",
            "debt_interest_mode": "subsequent_collection",
            "pre_deducted_interest": 0.0,
            "net_disbursed_amount": fresh_principal,
            "original_interest_amount": 0.0,
            "waived_interest_amount": 0.0,
            "waiver_reason": "",
            "waiver_date": None,
            "parent_loan_id": "",
            "loan_chain_id": "",
            "chain_start_date": to_datetime(row[6]),
        })
    if skipped_loans:
        print(f"  ⚠ Skipped {skipped_loans} loan rows with missing ID")

    # ---- Payments ----
    _, payment_rows_raw = iter_nonempty_rows(wb[payment_sheet])
    payments: List[Dict[str, Any]] = []
    skipped_payments = 0
    for row_number, row in payment_rows_raw:
        payment_id = str(row[0] or "").strip()
        loan_id = str(row[1] or "").strip()
        customer_id = str(row[2] or "").strip()
        if not payment_id:
            skipped_payments += 1
            continue
        payment_type = str(row[6] or "INTEREST").strip().upper() or "INTEREST"
        if payment_type not in VALID_PAYMENT_TYPES:
            print(f"  ⚠ Payment {payment_id}: Unknown type '{payment_type}' → recovered to INTEREST")
            payment_type = "INTEREST"
        amount = round(to_float(row[5]), 2)

        # Derive principal/interest based on type
        principal_amount = 0.0
        interest_amount = 0.0
        is_virtual = payment_type == "BALANCE"
        if payment_type in {"PRINCIPAL", "BALANCE"}:
            principal_amount = amount
        elif payment_type == "INTEREST":
            interest_amount = amount
        elif payment_type == "BOTH":
            # BOTH: the full amount is principal (interest portion is 0 in source data)
            principal_amount = amount

        payments.append({
            "payment_id": payment_id,
            "loan_id": loan_id,
            "customer_id": customer_id,
            "payment_date": to_datetime(row[4]),
            "amount": amount,
            "payment_type": payment_type,
            "payment_method": "CASH",
            "reference_number": "",
            "created_date": to_datetime(row[8]) if len(row) > 8 else None
                            or to_datetime(row[4]) or datetime.now(),
            "created_by": "SYSTEM",
            "notes": str((row[9] if len(row) > 9 else None)
                         or (row[7] if len(row) > 7 else None) or "").strip(),
            "principal_amount": principal_amount,
            "interest_amount": interest_amount,
            "help_category": "None",
            "is_virtual": is_virtual,
            "linked_successor_loan_id": "",
        })
    if skipped_payments:
        print(f"  ⚠ Skipped {skipped_payments} payment rows with missing ID")

    wb.close()
    return {"customers": customers, "loans": loans, "payments": payments}


# ---------------------------------------------------------------------------
# Deduplicate loan IDs (same logic as data_migration/migrate_data.py)
# ---------------------------------------------------------------------------

def deduplicate_loans(records: Dict[str, List[Dict[str, Any]]]) -> None:
    """Resolve duplicate loan IDs by reassigning new IDs to duplicates.
    Also patches related payments to point at the new loan ID."""
    import collections

    loans = records["loans"]
    payments = records["payments"]
    loan_groups: Dict[str, List[Dict[str, Any]]] = collections.defaultdict(list)
    for loan in loans:
        loan_groups[loan["loan_id"]].append(loan)

    # Find the highest numeric suffix to generate unique replacements
    numeric_ids = []
    for loan_id in loan_groups:
        match = re.search(r"(\d+)", loan_id)
        if match:
            numeric_ids.append(int(match.group(1)))
    next_numeric = max(numeric_ids or [0]) + 1

    reassigned: Dict[Tuple[str, str], str] = {}  # (old_loan_id, customer_id) → new_loan_id

    for loan_id, group in loan_groups.items():
        if len(group) <= 1:
            continue
        group.sort(key=lambda item: (str(item.get("start_date") or ""), item["customer_id"]))
        # Keep the first, reassign duplicates
        for duplicate in group[1:]:
            new_id = f"LN{next_numeric:03d}"
            next_numeric += 1
            reassigned[(loan_id, duplicate["customer_id"])] = new_id
            print(f"  ↔ Duplicate loan {loan_id} (customer {duplicate['customer_id']}) → {new_id}")
            duplicate["loan_id"] = new_id

    # Relink payments
    for payment in payments:
        key = (payment["loan_id"], payment["customer_id"])
        if key in reassigned:
            payment["loan_id"] = reassigned[key]


# ---------------------------------------------------------------------------
# Validate referential integrity
# ---------------------------------------------------------------------------

def validate_records(records: Dict[str, List[Dict[str, Any]]]) -> List[str]:
    """Return a list of validation warnings/errors."""
    issues = []
    customer_ids = {c["customer_id"] for c in records["customers"]}
    loan_ids = {l["loan_id"] for l in records["loans"]}

    # Check for duplicate primary keys
    import collections
    for entity, key in [("customers", "customer_id"), ("loans", "loan_id"), ("payments", "payment_id")]:
        counts = collections.Counter(r[key] for r in records[entity])
        dupes = {k: v for k, v in counts.items() if v > 1}
        if dupes:
            issues.append(f"DUPLICATE {entity} IDs: {dupes}")

    # Orphan loans (customer_id not in borrower_master)
    for loan in records["loans"]:
        if loan["customer_id"] not in customer_ids:
            issues.append(f"Loan {loan['loan_id']} references missing customer {loan['customer_id']}")

    # Orphan payments (loan_id not in loan_master)
    orphan_count = 0
    for pmt in records["payments"]:
        if pmt["loan_id"] not in loan_ids:
            orphan_count += 1
    if orphan_count:
        issues.append(f"{orphan_count} payments reference non-existent loan IDs")

    return issues


# ---------------------------------------------------------------------------
# Write to PostgreSQL
# ---------------------------------------------------------------------------

def wipe_and_load(engine, records: Dict[str, List[Dict[str, Any]]]) -> Dict[str, int]:
    """Truncate all tables and insert fresh data from the source records."""

    with engine.begin() as conn:
        # ---- WIPE (order matters: children before parents to avoid FK issues) ----
        print("\n  Truncating tables...")
        for table in [
            "audit_log", "help_records", "payment_transactions",
            "capital_injections", "loan_master", "borrower_master",
            "system_config"
        ]:
            conn.execute(text(f"TRUNCATE TABLE {table} CASCADE"))
        print("  ✓ All tables truncated")

        # ---- INSERT CUSTOMERS ----
        print(f"  Inserting {len(records['customers'])} customers...")
        for batch_start in range(0, len(records["customers"]), 500):
            batch = records["customers"][batch_start:batch_start + 500]
            conn.execute(
                text("""
                    INSERT INTO borrower_master
                    (customer_id, name, phone, email, address, id_proof_type,
                     id_proof_number, status, created_date, notes)
                    VALUES (:customer_id, :name, :phone, :email, :address,
                            :id_proof_type, :id_proof_number, :status,
                            :created_date, :notes)
                    ON CONFLICT (customer_id) DO NOTHING
                """),
                batch
            )
        print(f"  ✓ Customers inserted")

        # ---- INSERT LOANS ----
        print(f"  Inserting {len(records['loans'])} loans...")
        for batch_start in range(0, len(records["loans"]), 500):
            batch = records["loans"][batch_start:batch_start + 500]
            conn.execute(
                text("""
                    INSERT INTO loan_master
                    (loan_id, customer_id, principal_amount, add_on_principal,
                     fresh_principal,
                     interest_rate, loan_type, start_date, tenure_months, status,
                     fund_source, created_date, closed_date, notes,
                     transaction_type, debt_interest_mode, pre_deducted_interest,
                     net_disbursed_amount, original_interest_amount,
                     waived_interest_amount, waiver_reason, waiver_date,
                     parent_loan_id, loan_chain_id, chain_start_date)
                    VALUES (:loan_id, :customer_id, :principal_amount,
                            :add_on_principal, :fresh_principal, :interest_rate, :loan_type,
                            :start_date, :tenure_months, :status, :fund_source,
                            :created_date, :closed_date, :notes,
                            :transaction_type, :debt_interest_mode,
                            :pre_deducted_interest, :net_disbursed_amount,
                            :original_interest_amount, :waived_interest_amount,
                            :waiver_reason, :waiver_date,
                            :parent_loan_id, :loan_chain_id, :chain_start_date)
                    ON CONFLICT (loan_id) DO NOTHING
                """),
                batch
            )
        print(f"  ✓ Loans inserted")

        # ---- INSERT PAYMENTS ----
        print(f"  Inserting {len(records['payments'])} payments...")
        for batch_start in range(0, len(records["payments"]), 500):
            batch = records["payments"][batch_start:batch_start + 500]
            conn.execute(
                text("""
                    INSERT INTO payment_transactions
                    (payment_id, loan_id, customer_id, payment_date, amount,
                     payment_type, payment_method, reference_number,
                     created_date, created_by, notes,
                     principal_amount, interest_amount, help_category,
                     is_virtual, linked_successor_loan_id)
                    VALUES (:payment_id, :loan_id, :customer_id, :payment_date,
                            :amount, :payment_type, :payment_method,
                            :reference_number, :created_date, :created_by,
                            :notes, :principal_amount, :interest_amount,
                            :help_category, :is_virtual, :linked_successor_loan_id)
                    ON CONFLICT (payment_id) DO NOTHING
                """),
                batch
            )
        print(f"  ✓ Payments inserted")

        # ---- RESET SYSTEM CONFIG SEQUENCES ----
        print("  Resetting ID sequences in system_config...")
        # Find max numeric suffix for each entity type
        max_customer = 0
        for c in records["customers"]:
            match = re.search(r"(\d+)", c["customer_id"])
            if match:
                max_customer = max(max_customer, int(match.group(1)))

        max_loan = 0
        for l in records["loans"]:
            match = re.search(r"(\d+)", l["loan_id"])
            if match:
                max_loan = max(max_loan, int(match.group(1)))

        max_payment = 0
        for p in records["payments"]:
            match = re.search(r"(\d+)", p["payment_id"])
            if match:
                max_payment = max(max_payment, int(match.group(1)))

        configs = [
            ("next_customer_id", str(max_customer + 1), "Next customer ID"),
            ("next_loan_id", str(max_loan + 1), "Next loan ID"),
            ("next_payment_id", str(max_payment + 1), "Next payment ID"),
            ("next_injection_id", "1", "Next capital injection ID"),
            ("next_audit_id", "1", "Next audit log ID"),
            ("next_help_id", "1", "Next help ID"),
            ("schema_version", "2026-03-28-postgres-v1", "Migration schema version"),
        ]
        now = datetime.now()
        for key, val, desc in configs:
            conn.execute(
                text("""
                    INSERT INTO system_config (config_key, config_value, description, last_updated)
                    VALUES (:key, :val, :desc, :ts)
                    ON CONFLICT (config_key) DO UPDATE
                    SET config_value = :val, description = :desc, last_updated = :ts
                """),
                {"key": key, "val": val, "desc": desc, "ts": now}
            )
        print(f"  ✓ Sequences reset: customer={max_customer+1}, loan={max_loan+1}, payment={max_payment+1}")

    # ---- POST-INSERT VERIFICATION ----
    counts = {}
    with engine.connect() as conn:
        for table, entity in [
            ("borrower_master", "customers"),
            ("loan_master", "loans"),
            ("payment_transactions", "payments"),
        ]:
            result = conn.execute(text(f"SELECT COUNT(*) FROM {table}"))
            counts[entity] = result.scalar()

    return counts


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Migrate DIAMOND FINANCE DATA.xlsm → PostgreSQL")
    parser.add_argument("--source", type=str, default=str(SOURCE_DEFAULT),
                        help="Path to source XLSM file")
    parser.add_argument("--dry-run", action="store_true",
                        help="Validate source data without writing to database")
    args = parser.parse_args()

    source_file = Path(args.source)
    if not source_file.exists():
        print(f"✗ Source file not found: {source_file}")
        sys.exit(1)

    database_url = os.environ.get("DATABASE_URL", "")
    if not database_url and not args.dry_run:
        print("✗ DATABASE_URL environment variable is required")
        sys.exit(1)

    print("=" * 60)
    print("  VAIRAM FINCORP — Excel → PostgreSQL Migration")
    print("=" * 60)
    print(f"  Source: {source_file}")
    print(f"  Mode:   {'DRY RUN (no database writes)' if args.dry_run else 'LIVE'}")
    print()

    # Step 1: Extract
    print("[1/4] Extracting records from Excel...")
    records = load_source_records(source_file)
    print(f"  Extracted: {len(records['customers'])} customers, "
          f"{len(records['loans'])} loans, {len(records['payments'])} payments")

    # Step 2: Deduplicate
    print("\n[2/4] Deduplicating loan IDs...")
    deduplicate_loans(records)

    # Step 3: Validate
    print("\n[3/4] Validating referential integrity...")
    issues = validate_records(records)
    if issues:
        print(f"  ⚠ {len(issues)} issues found:")
        for issue in issues:
            print(f"    - {issue}")
    else:
        print("  ✓ All records pass validation")

    if args.dry_run:
        print("\n[4/4] DRY RUN — skipping database writes")
        print("\n" + "=" * 60)
        print("  Dry run complete. No data was written.")
        print("=" * 60)
        return

    # Step 4: Load into PostgreSQL
    print("\n[4/4] Loading into PostgreSQL...")
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    engine = create_engine(database_url, pool_size=2, max_overflow=5, pool_timeout=30)

    # Ensure schema exists before loading data
    from main import PostgresDB
    pg = PostgresDB(database_url)
    pg._ensure_schema()
    print("  ✓ Schema verified")

    db_counts = wipe_and_load(engine, records)

    print("\n" + "=" * 60)
    print("  MIGRATION COMPLETE")
    print("=" * 60)
    print(f"  Customers in DB: {db_counts['customers']}  (source: {len(records['customers'])})")
    print(f"  Loans in DB:     {db_counts['loans']}  (source: {len(records['loans'])})")
    print(f"  Payments in DB:  {db_counts['payments']}  (source: {len(records['payments'])})")

    # Verify counts match
    all_match = all(
        db_counts[entity] == len(records[entity])
        for entity in ["customers", "loans", "payments"]
    )
    if all_match:
        print("\n  ✓ All counts match — migration is verified!")
    else:
        print("\n  ✗ COUNT MISMATCH — review the output above for data issues")

    engine.dispose()


if __name__ == "__main__":
    main()
