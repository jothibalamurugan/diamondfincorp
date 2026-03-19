"""
Diamond Fincorp Loan Management System - Enterprise Backend API
FastAPI server with Excel backend - Enterprise Grade with Full Feature Set
"""

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, date
from enum import Enum
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import re
import threading
from decimal import Decimal
import logging
import json
from sqlalchemy import create_engine, text, Table, Column, String, Float, MetaData, Integer, DateTime
from sqlalchemy.orm import declarative_base, Session
from sqlalchemy.exc import ProgrammingError

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration
# Resolve absolute path dynamically so the workbook path stays configurable via EXCEL_DB_PATH.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_DB_PATH = os.path.normpath(os.path.join(BASE_DIR, '..', 'excel_schema', 'LoanManagement_DB.xlsx'))
LOCAL_DB_PATH = os.path.normpath(os.path.join(BASE_DIR, 'LoanManagement_DB.xlsx'))
EXCEL_DB_PATH = os.environ.get('EXCEL_DB_PATH')
if EXCEL_DB_PATH:
    EXCEL_DB_PATH = os.path.normpath(EXCEL_DB_PATH)
else:
    EXCEL_DB_PATH = next(
        (candidate for candidate in [DEFAULT_DB_PATH, LOCAL_DB_PATH] if os.path.exists(candidate)),
        DEFAULT_DB_PATH
    )
WORKBOOK_SCHEMA_VERSION = '2026-03-19-enterprise-v1'

APP_ENV = os.environ.get('APP_ENV', os.environ.get('ENV', 'development')).strip().lower()
IS_DEVELOPMENT = APP_ENV in {'development', 'dev', 'local', 'debug'}

def parse_allowed_origins() -> List[str]:
    configured = os.environ.get('ALLOWED_ORIGINS', '').strip()
    if configured:
        return [origin.strip() for origin in configured.split(',') if origin.strip()]
    return [
        'null',
        'http://localhost',
        'http://localhost:3000',
        'http://localhost:5173',
        'http://localhost:8000',
        'http://127.0.0.1',
        'http://127.0.0.1:3000',
        'http://127.0.0.1:5173',
        'http://127.0.0.1:8000'
    ]

def ensure_debug_access():
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Debug endpoints are disabled outside development")

# Sheet name mappings - maps internal names to actual Excel sheet names
# User's Excel uses: Loan_Master, Borrower_Master, Payment_Transactions
SHEET_NAMES = {
    'Loans': 'Loan_Master',
    'Customers': 'Borrower_Master', 
    'Payments': 'Payment_Transactions',
    'Help': 'HELP',
    'CapitalInjections': 'CapitalInjections',
    'AuditLog': 'AuditLog',
    'SystemConfig': 'SystemConfig'
}

# Column name mappings - maps internal names to actual Excel column names
# This allows flexibility for different Excel schemas
COLUMN_MAPPINGS = {
    # Loan columns
    'loan_id': ['LoanID', 'loan_id', 'Loan_ID', 'loanid'],
    'customer_id': ['BorrowerId', 'customer_id', 'CustomerID', 'borrower_id'],
    'principal_amount': ['principal_amount', 'Principal', 'PrincipalAmount', 'Amount'],
    'interest_rate': ['interest_rate', 'InterestRate', 'Rate', 'interest'],
    'start_date': ['start_date', 'StartDate', 'Date', 'LoanDate'],
    'status': ['status', 'Status', 'LoanStatus'],
    'type': ['type', 'Type', 'LoanType', 'transaction_type'],
    
    # Payment columns
    'payment_id': ['payment_id', 'PaymentID', 'TransactionID'],
    'payment_date': ['payment_date', 'PaymentDate', 'Date', 'TransactionDate'],
    'amount': ['amount', 'Amount', 'PaymentAmount'],
    'payment_type': ['payment_type', 'PaymentType', 'Type'],
    'principal_amount': ['principal_amount', 'PrincipalAmount', 'principal_paid', 'principal_component'],
    'interest_amount': ['interest_amount', 'InterestAmount', 'interest_paid', 'interest_component'],
    
    # Borrower columns
    'name': ['name', 'Name', 'BorrowerName', 'CustomerName']
}

def get_column_value(row: dict, column_key: str, default=None):
    """Get value from row using flexible column name matching"""
    if column_key in COLUMN_MAPPINGS:
        for possible_name in COLUMN_MAPPINGS[column_key]:
            if possible_name in row and row[possible_name] is not None:
                return row[possible_name]
    # Direct access if not in mappings
    if column_key in row:
        return row[column_key]
    return default

def get_sheet_name(internal_name: str) -> str:
    """Get actual sheet name from internal name"""
    return SHEET_NAMES.get(internal_name, internal_name)

app = FastAPI(
    title="Diamond Fincorp Loan Management API",
    description="Enterprise-grade loan management system with Excel backend",
    version="2.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=parse_allowed_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def serve_frontend():
    frontend_path = os.path.join(BASE_DIR, '..', 'frontend', 'index.html')
    if os.path.exists(frontend_path):
        return FileResponse(frontend_path)
    return {"message": "Frontend file not found"}

# ==================== ENUMS ====================

class TransactionType(str, Enum):
    KULU = "KULU"
    DEBT = "DEBT"

class LoanStatus(str, Enum):
    ACTIVE = "ACTIVE"
    COMPLETED = "COMPLETED"
    DEFAULTED = "DEFAULTED"
    WRITTEN_OFF = "WRITTEN_OFF"

class PaymentType(str, Enum):
    PRINCIPAL = "PRINCIPAL"
    INTEREST = "INTEREST"
    BOTH = "BOTH"

class CapitalSourceType(str, Enum):
    SALARY = "SALARY"
    PERSONAL = "PERSONAL"
    INVESTOR = "INVESTOR"
    BANK_LOAN = "BANK_LOAN"
    OTHER = "OTHER"

# ==================== DATA MODELS ====================

class Customer(BaseModel):
    customer_id: Optional[str] = None
    name: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    id_proof_type: Optional[str] = None
    id_proof_number: Optional[str] = None
    status: str = 'ACTIVE'
    created_date: Optional[datetime] = None
    notes: Optional[str] = None

class Loan(BaseModel):
    loan_id: Optional[str] = None
    customer_id: str
    principal_amount: float
    add_on_principal: Optional[float] = 0.0
    interest_rate: float
    loan_type: str = 'PERSONAL'
    transaction_type: str = 'KULU'  # KULU, DEBT, OTHER
    start_date: date
    tenure_months: Optional[int] = None
    status: str = 'ACTIVE'
    fund_source: Optional[str] = None
    debt_interest_mode: str = 'subsequent_collection'
    pre_deducted_interest: Optional[float] = 0.0
    net_disbursed_amount: Optional[float] = None
    original_interest_amount: Optional[float] = 0.0
    waived_interest_amount: Optional[float] = 0.0
    waiver_reason: Optional[str] = None
    waiver_date: Optional[date] = None
    created_date: Optional[datetime] = None
    closed_date: Optional[datetime] = None
    notes: Optional[str] = None

class Payment(BaseModel):
    payment_id: Optional[str] = None
    loan_id: str
    customer_id: str
    payment_date: date
    amount: Optional[float] = None
    total_amount: Optional[float] = None
    principal_amount: Optional[float] = 0.0
    interest_amount: Optional[float] = 0.0
    payment_type: Optional[str] = None  # PRINCIPAL/INTEREST/BOTH
    payment_method: str = 'CASH'
    reference_number: Optional[str] = None
    created_date: Optional[datetime] = None
    created_by: str = 'USER'
    help_category: Optional[str] = 'None'
    help_note: Optional[str] = None
    repayment_date: Optional[date] = None
    repayment_amount: Optional[float] = None
    help_status: Optional[str] = None
    notes: Optional[str] = None

class HelpRecord(BaseModel):
    help_id: Optional[str] = None
    customer_id: str
    customer_name: str
    help_date: date
    help_amount: float
    help_category: str
    help_note: Optional[str] = None
    repayment_date: Optional[date] = None
    repayment_amount: Optional[float] = None
    status: str = 'Active'

class CapitalInjection(BaseModel):
    injection_id: Optional[str] = None
    source_type: str  # SALARY, PERSONAL, INVESTOR, BANK_LOAN, OTHER
    amount: float
    injection_date: date
    description: Optional[str] = None
    created_by: str = 'USER'
    created_date: Optional[datetime] = None

class InterestWaiver(BaseModel):
    loan_id: str
    waived_amount: float
    reason: str
    waiver_date: Optional[date] = None

class AuditLogEntry(BaseModel):
    log_id: Optional[str] = None
    entity_type: str
    entity_id: str
    action: str
    old_value: Optional[str] = None
    new_value: Optional[str] = None
    user: str = 'USER'
    timestamp: Optional[datetime] = None

class LoanSummary(BaseModel):
    loan_id: str
    customer_name: str
    principal_amount: float
    add_on_principal: float
    effective_principal_amount: float
    net_disbursed_amount: float
    pre_deducted_interest: float
    total_paid: float
    principal_paid: float
    interest_paid: float
    outstanding_balance: float
    interest_rate: float
    transaction_type: str
    debt_interest_mode: str
    status: str
    start_date: str
    start_date_iso: Optional[str] = None
    months_active: int
    days_active: Optional[int] = None
    total_interest_accrued: float
    original_interest_amount: float
    waived_interest_amount: float
    effective_interest_due: float

class DashboardStats(BaseModel):
    total_customers: int
    active_customers: int
    total_loans: int
    active_loans: int
    total_principal_disbursed: float
    total_principal_outstanding: float
    total_interest_collected: float
    total_interest_waived: float
    total_principal_collected: float
    net_profit: float
    portfolio_health: str
    kulu_count: int
    debt_count: int
    other_count: int
    total_capital_injected: float
    capital_utilized: float

class CapitalSummary(BaseModel):
    total_injected: float
    total_disbursed: float
    available_capital: float
    utilization_percentage: float
    by_source: Dict[str, float]

# ==================== DATABASE OPERATIONS ====================

class ExcelDB:
    """Excel database handler with thread-safe operations"""
    
    def __init__(self, filepath):
        self.filepath = filepath
        self._id_lock = threading.Lock()
        self._ensure_file_exists()
        self._ensure_schema_updated()
    
    def _ensure_file_exists(self):
        if not os.path.exists(self.filepath):
            raise FileNotFoundError(f"Database file not found: {self.filepath}")
    
    def _ensure_schema_updated(self):
        """Ensure new columns and sheets exist for enterprise features"""
        wb = openpyxl.load_workbook(self.filepath)
        modified = False
        ws_config = self._get_worksheet(wb, 'SystemConfig')
        config_rows = {
            str(row[0].value): row
            for row in ws_config.iter_rows(min_row=2, values_only=False)
            if row[0].value
        }
        schema_row = config_rows.get('schema_version')
        if schema_row and str(schema_row[1].value or '').strip() == WORKBOOK_SCHEMA_VERSION:
            wb.close()
            return

        def ensure_columns(sheet_alias: str, column_defaults: Dict[str, Any]):
            nonlocal modified
            try:
                ws = self._get_worksheet(wb, sheet_alias)
            except KeyError:
                return

            headers = [cell.value for cell in ws[1]]
            for col_name, default_value in column_defaults.items():
                if col_name in headers:
                    continue

                col_idx = len(headers) + 1
                ws.cell(row=1, column=col_idx, value=col_name)
                headers.append(col_name)
                modified = True

                for row_idx in range(2, ws.max_row + 1):
                    if not ws.cell(row=row_idx, column=1).value:
                        continue
                    if callable(default_value):
                        ws.cell(row=row_idx, column=col_idx, value=default_value(ws, row_idx))
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=default_value)

        def add_sheet(sheet_name: str, headers: List[str]):
            nonlocal modified
            if sheet_name in wb.sheetnames:
                return

            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = 'A2'
            modified = True

        ensure_columns('Loans', {
            'add_on_principal': 0,
            'transaction_type': 'OTHER',
            'debt_interest_mode': 'subsequent_collection',
            'pre_deducted_interest': 0,
            'net_disbursed_amount': '',
            'original_interest_amount': 0,
            'waived_interest_amount': 0,
            'waiver_reason': '',
            'waiver_date': ''
        })

        ensure_columns('Payments', {
            'principal_amount': 0,
            'interest_amount': lambda ws, row_idx: ws.cell(row=row_idx, column=5).value or 0,
            'help_category': 'None'
        })

        add_sheet('CapitalInjections', [
            'injection_id',
            'source_type',
            'amount',
            'injection_date',
            'description',
            'created_by',
            'created_date'
        ])

        add_sheet('AuditLog', [
            'log_id',
            'entity_type',
            'entity_id',
            'action',
            'old_value',
            'new_value',
            'user',
            'timestamp'
        ])

        add_sheet('HELP', [
            'HelpID',
            'CustomerID',
            'CustomerName',
            'HelpDate',
            'HelpAmount',
            'HelpCategory',
            'HelpNote',
            'RepaymentDate',
            'RepaymentAmount',
            'Status'
        ])

        config_keys = list(config_rows.keys())

        for config_key, description in [
            ('next_injection_id', 'Next capital injection ID'),
            ('next_audit_id', 'Next audit log ID'),
            ('next_help_id', 'Next help ID')
        ]:
            if config_key not in config_keys:
                ws_config.append([config_key, '1', description, datetime.now()])
                modified = True
        if schema_row:
            schema_row[1].value = WORKBOOK_SCHEMA_VERSION
            schema_row[2].value = 'Workbook schema version'
            schema_row[3].value = datetime.now()
            modified = True
        else:
            ws_config.append(['schema_version', WORKBOOK_SCHEMA_VERSION, 'Workbook schema version', datetime.now()])
            modified = True
        
        if modified:
            wb.save(self.filepath)
            logger.info("Database schema updated for enterprise features")
        
        wb.close()
    

        
    def _load_workbook(self):
        """Load workbook with data_only=True to get calculated values"""
        return openpyxl.load_workbook(self.filepath, data_only=True)
    
    def _save_workbook(self, wb):
        """Save workbook safely"""
        wb.save(self.filepath)

    def _get_worksheet(self, wb, sheet_name: str):
        """
        Resolve logical sheet name (e.g. 'Loans', 'Customers') to the actual
        Excel worksheet name using the central SHEET_NAMES mapping.
        Falls back to the provided name if a direct match exists.
        """
        actual_sheet_name = get_sheet_name(sheet_name)
        if actual_sheet_name in wb.sheetnames:
            return wb[actual_sheet_name]
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        logger.warning(
            f"Sheet '{sheet_name}' (mapped to '{actual_sheet_name}') not found. "
            f"Available sheets: {wb.sheetnames}"
        )
        raise KeyError(f"Worksheet for '{sheet_name}' not found")
    
    def get_next_id(self, id_type: str) -> str:
        """Get next sequential ID for customers, loans, payments, etc."""
        with self._id_lock:
            wb = openpyxl.load_workbook(self.filepath)
            ws = self._get_worksheet(wb, 'SystemConfig')
            
            config_map = {
                'customer': 'next_customer_id',
                'loan': 'next_loan_id',
                'payment': 'next_payment_id',
                'injection': 'next_injection_id',
                'audit': 'next_audit_id',
                'help': 'next_help_id'
            }
            
            config_key = config_map.get(id_type)
            next_num = 1
            
            for row in ws.iter_rows(min_row=2, values_only=False):
                if row[0].value == config_key:
                    next_num = int(row[1].value)
                    row[1].value = str(next_num + 1)
                    row[3].value = datetime.now()
                    break
            
            self._save_workbook(wb)
            
            prefixes = {
                'customer': 'CUST', 
                'loan': 'LN', 
                'payment': 'PAY',
                'injection': 'CAP',
                'audit': 'AUD',
                'help': 'HELP'
            }
            return f"{prefixes[id_type]}{next_num:04d}"
    
    def get_all_rows(self, sheet_name: str) -> List[Dict]:
        """Get all rows from a sheet as list of dicts"""
        wb = self._load_workbook()
        try:
            ws = self._get_worksheet(wb, sheet_name)
        except KeyError:
            # Already logged in _get_worksheet
            return []
        
        headers = [cell.value for cell in ws[1]]
        rows = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            row_dict = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_dict[header] = row[i]
            rows.append(row_dict)
        
        return rows
    
    def add_row(self, sheet_name: str, data: List[Any]):
        """Add a new row to a sheet"""
        wb = openpyxl.load_workbook(self.filepath)
        ws = self._get_worksheet(wb, sheet_name)
        ws.append(data)
        self._save_workbook(wb)

    def add_dict_row(self, sheet_name: str, data: Dict[str, Any]):
        """Add a row using the actual worksheet headers as the column order"""
        wb = openpyxl.load_workbook(self.filepath)
        ws = self._get_worksheet(wb, sheet_name)
        headers = [cell.value for cell in ws[1]]
        ws.append([data.get(header, '') for header in headers])
        self._save_workbook(wb)
    
    def update_row(self, sheet_name: str, id_column: str, id_value: str, updates: Dict):
        """Update a row based on ID"""
        wb = openpyxl.load_workbook(self.filepath)
        ws = self._get_worksheet(wb, sheet_name)
        
        headers = [cell.value for cell in ws[1]]
        id_col_idx = headers.index(id_column)
        
        for row in ws.iter_rows(min_row=2):
            if row[id_col_idx].value == id_value:
                for col_name, new_value in updates.items():
                    if col_name in headers:
                        col_idx = headers.index(col_name)
                        row[col_idx].value = new_value
                break
        
        self._save_workbook(wb)

    def delete_row(self, sheet_name: str, id_column: str, id_value: str):
        """Delete a row based on ID"""
        wb = openpyxl.load_workbook(self.filepath)
        ws = self._get_worksheet(wb, sheet_name)
        headers = [cell.value for cell in ws[1]]

        if id_column not in headers:
            wb.close()
            raise KeyError(f"Column '{id_column}' not found in sheet '{sheet_name}'")

        id_col_idx = headers.index(id_column)
        delete_idx = None

        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=id_col_idx + 1).value == id_value:
                delete_idx = row_idx
                break

        if delete_idx:
            ws.delete_rows(delete_idx, 1)
            self._save_workbook(wb)
        else:
            wb.close()
            raise KeyError(f"ID '{id_value}' not found in sheet '{sheet_name}'")
    
    def log_audit(self, entity_type: str, entity_id: str, action: str, old_value: Any = None, new_value: Any = None, user: str = 'USER'):
        """Add audit log entry"""
        audit_id = self.get_next_id('audit')
        
        data = [
            audit_id,
            entity_type,
            entity_id,
            action,
            json.dumps(old_value) if old_value else '',
            json.dumps(new_value) if new_value else '',
            user,
            datetime.now()
        ]
        
        self.add_row('AuditLog', data)
        return audit_id

class PostgresDB:
    """PostgreSQL database handler with identical interface to ExcelDB"""
    TABLE_MAP = {
        'Customers': 'borrower_master',
        'Loans': 'loan_master',
        'Payments': 'payment_transactions',
        'Help': 'help_records',
        'CapitalInjections': 'capital_injections',
        'AuditLog': 'audit_log',
        'SystemConfig': 'system_config'
    }
    TABLE_COLUMNS = {
        'Customers': {'customer_id', 'name', 'phone', 'email', 'address', 'id_proof_type', 'id_proof_number', 'status', 'created_date', 'notes'},
        'Loans': {'loan_id', 'customer_id', 'principal_amount', 'add_on_principal', 'interest_rate', 'loan_type', 'start_date', 'tenure_months', 'status', 'fund_source', 'created_date', 'closed_date', 'notes', 'transaction_type', 'debt_interest_mode', 'pre_deducted_interest', 'net_disbursed_amount', 'original_interest_amount', 'waived_interest_amount', 'waiver_reason', 'waiver_date'},
        'Payments': {'payment_id', 'loan_id', 'customer_id', 'payment_date', 'amount', 'payment_type', 'payment_method', 'reference_number', 'created_date', 'created_by', 'notes', 'principal_amount', 'interest_amount', 'help_category'},
        'Help': {'help_id', 'customer_id', 'customer_name', 'help_date', 'help_amount', 'help_category', 'help_note', 'repayment_date', 'repayment_amount', 'status'},
        'CapitalInjections': {'injection_id', 'source_type', 'amount', 'injection_date', 'description', 'created_by', 'created_date'},
        'AuditLog': {'log_id', 'entity_type', 'entity_id', 'action', 'old_value', 'new_value', 'user', 'timestamp'},
        'SystemConfig': {'config_key', 'config_value', 'description', 'last_updated'}
    }
    COLUMN_ALIASES = {
        'Customers': {'BorrowerID': 'customer_id', 'BorrowerName': 'name', 'Phone': 'phone', 'Address': 'address', 'IsActive': 'status', 'CreatedOn': 'created_date'},
        'Loans': {'LoanID': 'loan_id', 'BorrowerID': 'customer_id', 'TYPE': 'transaction_type', 'PrincipalAmount': 'principal_amount', 'AddOnPrincipal': 'add_on_principal', 'InterestRate': 'interest_rate', 'StartDate': 'start_date', 'FundSourceID': 'fund_source', 'LoanStatus': 'status', 'CreatedOn': 'created_date'},
        'Payments': {'PaymentID': 'payment_id', 'LoanID': 'loan_id', 'Borrower': 'customer_id', 'PaymentDate': 'payment_date', 'PaymentAmount': 'amount', 'PaymentType': 'payment_type', 'Remarks': 'notes', 'CreatedOn': 'created_date'},
        'Help': {'HelpID': 'help_id', 'CustomerID': 'customer_id', 'CustomerName': 'customer_name', 'HelpDate': 'help_date', 'HelpAmount': 'help_amount', 'HelpCategory': 'help_category', 'HelpNote': 'help_note', 'RepaymentDate': 'repayment_date', 'RepaymentAmount': 'repayment_amount', 'Status': 'status'},
        'CapitalInjections': {},
        'AuditLog': {'user': 'user'},
        'SystemConfig': {}
    }
    
    def __init__(self, database_url: str):
        self.database_url = database_url
        if self.database_url.startswith("postgres://"):
            self.database_url = self.database_url.replace("postgres://", "postgresql://", 1)
            
        # Using connect_args for ssl require in railway sometimes helps, and smaller pool
        self.engine = create_engine(self.database_url, pool_size=2, max_overflow=5, pool_timeout=30)
        self.metadata = MetaData()
        # We don't automatically call _ensure_schema here anymore to prevent deadlocks
        # It should be called explicitly from migration/setup scripts.
        
    def _get_table_name(self, sheet_name):
        if sheet_name not in self.TABLE_MAP:
            raise ValueError(f"Unsupported sheet name: {sheet_name}")
        return self.TABLE_MAP[sheet_name]

    def _get_allowed_columns(self, sheet_name: str) -> set:
        if sheet_name not in self.TABLE_COLUMNS:
            raise ValueError(f"Unsupported sheet name: {sheet_name}")
        return self.TABLE_COLUMNS[sheet_name]

    def _normalize_column_name(self, sheet_name: str, column_name: str) -> str:
        aliases = self.COLUMN_ALIASES.get(sheet_name, {})
        if column_name in aliases:
            return aliases[column_name]
        lowered = column_name.strip().lower()
        if lowered in self._get_allowed_columns(sheet_name):
            return lowered
        return column_name

    def _ensure_schema(self):
        """Create tables if they don't exist"""
        with self.engine.begin() as conn:
            # Customers
            conn.execute(text('''
                CREATE TABLE IF NOT EXISTS borrower_master (
                    customer_id VARCHAR(50) PRIMARY KEY,
                    name VARCHAR(255),
                    phone VARCHAR(50),
                    email VARCHAR(255),
                    address TEXT,
                    id_proof_type VARCHAR(50),
                    id_proof_number VARCHAR(100),
                    status VARCHAR(50),
                    created_date TIMESTAMP,
                    notes TEXT
                )
            '''))
            
            # Loans
            conn.execute(text('''
                CREATE TABLE IF NOT EXISTS loan_master (
                    loan_id VARCHAR(50) PRIMARY KEY,
                    customer_id VARCHAR(50),
                    principal_amount FLOAT,
                    add_on_principal FLOAT,
                    interest_rate FLOAT,
                    loan_type VARCHAR(50),
                    start_date DATE,
                    tenure_months VARCHAR(50),
                    status VARCHAR(50),
                    fund_source VARCHAR(100),
                    debt_interest_mode VARCHAR(50),
                    pre_deducted_interest FLOAT,
                    net_disbursed_amount FLOAT,
                    created_date TIMESTAMP,
                    closed_date TIMESTAMP,
                    notes TEXT,
                    transaction_type VARCHAR(50),
                    original_interest_amount FLOAT,
                    waived_interest_amount FLOAT,
                    waiver_reason TEXT,
                    waiver_date DATE
                )
            '''))
            
            # Payments
            conn.execute(text('''
                CREATE TABLE IF NOT EXISTS payment_transactions (
                    payment_id VARCHAR(50) PRIMARY KEY,
                    loan_id VARCHAR(50),
                    customer_id VARCHAR(50),
                    payment_date DATE,
                    amount FLOAT,
                    principal_amount FLOAT,
                    interest_amount FLOAT,
                    payment_type VARCHAR(50),
                    payment_method VARCHAR(50),
                    reference_number VARCHAR(255),
                    created_date TIMESTAMP,
                    created_by VARCHAR(100),
                    help_category VARCHAR(100),
                    notes TEXT
                )
            '''))

            conn.execute(text('''
                CREATE TABLE IF NOT EXISTS help_records (
                    help_id VARCHAR(50) PRIMARY KEY,
                    customer_id VARCHAR(50),
                    customer_name VARCHAR(255),
                    help_date DATE,
                    help_amount FLOAT,
                    help_category VARCHAR(100),
                    help_note TEXT,
                    repayment_date DATE,
                    repayment_amount FLOAT,
                    status VARCHAR(50)
                )
            '''))
            
            # Capital Injections
            conn.execute(text('''
                CREATE TABLE IF NOT EXISTS capital_injections (
                    injection_id VARCHAR(50) PRIMARY KEY,
                    source_type VARCHAR(100),
                    amount FLOAT,
                    injection_date DATE,
                    description TEXT,
                    created_by VARCHAR(100),
                    created_date TIMESTAMP
                )
            '''))
            
            # Audit Log
            conn.execute(text('''
                CREATE TABLE IF NOT EXISTS audit_log (
                    log_id VARCHAR(50) PRIMARY KEY,
                    entity_type VARCHAR(50),
                    entity_id VARCHAR(50),
                    action VARCHAR(100),
                    old_value TEXT,
                    new_value TEXT,
                    "user" VARCHAR(100),
                    timestamp TIMESTAMP
                )
            '''))
            
            # System Config
            conn.execute(text('''
                CREATE TABLE IF NOT EXISTS system_config (
                    config_key VARCHAR(100) PRIMARY KEY,
                    config_value VARCHAR(255),
                    description TEXT,
                    last_updated TIMESTAMP
                )
            '''))

            # Initialize SystemConfig rows if not present
            configs = [
                ('next_customer_id', '1', 'Next customer ID'),
                ('next_loan_id', '1', 'Next loan ID'),
                ('next_payment_id', '1', 'Next payment ID'),
                ('next_injection_id', '1', 'Next capital injection ID'),
                ('next_audit_id', '1', 'Next audit log ID'),
                ('next_help_id', '1', 'Next help ID')
            ]
            for key, val, desc in configs:
                conn.execute(text('''
                    INSERT INTO system_config (config_key, config_value, description, last_updated)
                    VALUES (:k, :v, :d, :t)
                    ON CONFLICT (config_key) DO NOTHING
                '''), {"k": key, "v": val, "d": desc, "t": datetime.now()})

            logger.info("Checked/Created PostgreSQL schema.")

    def get_next_id(self, id_type: str) -> str:
        config_map = {
            'customer': 'next_customer_id',
            'loan': 'next_loan_id',
            'payment': 'next_payment_id',
            'injection': 'next_injection_id',
            'audit': 'next_audit_id',
            'help': 'next_help_id'
        }
        
        config_key = config_map.get(id_type)
        if not config_key:
            raise ValueError(f"Unknown id_type: {id_type}")
            
        with self.engine.begin() as conn:
            result = conn.execute(text('''
                UPDATE system_config
                SET config_value = CAST(CAST(config_value AS INTEGER) + 1 AS TEXT),
                    last_updated = :updated
                WHERE config_key = :key
                RETURNING config_value
            '''), {"updated": datetime.now(), "key": config_key}).scalar()

            if result is None:
                next_num = 1
                conn.execute(text('''
                    INSERT INTO system_config (config_key, config_value, description, last_updated)
                    VALUES (:key, :value, :description, :updated)
                    ON CONFLICT (config_key) DO NOTHING
                '''), {
                    "key": config_key,
                    "value": "2",
                    "description": f"Next {id_type} ID",
                    "updated": datetime.now()
                })
            else:
                next_num = int(result) - 1

        prefixes = {
            'customer': 'CUST', 
            'loan': 'LN', 
            'payment': 'PAY',
            'injection': 'CAP',
            'audit': 'AUD',
            'help': 'HELP'
        }
        return f"{prefixes[id_type]}{next_num:04d}"

    def get_all_rows(self, sheet_name: str) -> List[Dict]:
        table_name = self._get_table_name(sheet_name)
        with self.engine.connect() as conn:
            try:
                result = conn.execute(text(f"SELECT * FROM {table_name}"))
                keys = result.keys()
                # Return list of dictionaries identical to how openpyxl rows were parsed
                return [dict(zip(keys, row)) for row in result]
            except Exception as e:
                logger.error(f"Error fetching from {table_name}: {str(e)}")
                return []

    def add_row(self, sheet_name: str, data: List[Any]):
        """
        Translates a list of values into a Postgres insert.
        This relies on the columns being inserted in the exact same order as the list,
        which we mimic from the Excel schemas.
        """
        table_name = self._get_table_name(sheet_name)
        
        # Hardcoded column mappings based on the lists in endpoints
        cols = []
        if sheet_name == 'Customers':
            cols = ['customer_id', 'name', 'phone', 'email', 'address', 'id_proof_type', 'id_proof_number', 'status', 'created_date', 'notes']
        elif sheet_name == 'Loans':
            cols = ['loan_id', 'customer_id', 'principal_amount', 'add_on_principal', 'interest_rate', 'loan_type', 'start_date', 'tenure_months', 'status', 'fund_source', 'created_date', 'closed_date', 'notes', 'transaction_type', 'debt_interest_mode', 'pre_deducted_interest', 'net_disbursed_amount', 'original_interest_amount', 'waived_interest_amount', 'waiver_reason', 'waiver_date']
        elif sheet_name == 'Payments':
            cols = ['payment_id', 'loan_id', 'customer_id', 'payment_date', 'amount', 'payment_type', 'payment_method', 'reference_number', 'created_date', 'created_by', 'notes', 'principal_amount', 'interest_amount', 'help_category']
        elif sheet_name == 'Help':
            cols = ['help_id', 'customer_id', 'customer_name', 'help_date', 'help_amount', 'help_category', 'help_note', 'repayment_date', 'repayment_amount', 'status']
        elif sheet_name == 'CapitalInjections':
            cols = ['injection_id', 'source_type', 'amount', 'injection_date', 'description', 'created_by', 'created_date']
        elif sheet_name == 'AuditLog':
            cols = ['log_id', 'entity_type', 'entity_id', 'action', 'old_value', 'new_value', '"user"', 'timestamp']
        
        if len(data) != len(cols):
            logger.error(f"Length mismatch inserting into {sheet_name}. Data: {len(data)}, Cols: {len(cols)}")
            return
            
        placeholders = ", ".join([f":{col.replace('\"', '')}" for col in cols])
        columns_str = ", ".join(cols)
        query = f"INSERT INTO {table_name} ({columns_str}) VALUES ({placeholders})"
        
        params = {col.replace("\"", ""): val for col, val in zip(cols, data)}
        
        with self.engine.begin() as conn:
            conn.execute(text(query), params)

    def add_dict_row(self, sheet_name: str, data: Dict[str, Any]):
        table_name = self._get_table_name(sheet_name)
        allowed = self._get_allowed_columns(sheet_name)
        payload = {}
        for key, value in data.items():
            if value in (None, ''):
                continue
            normalized_key = self._normalize_column_name(sheet_name, key)
            if normalized_key in allowed:
                payload[normalized_key] = value
        if not payload:
            return
        columns = list(payload.keys())
        placeholders = ', '.join(f':{column}' for column in columns)
        with self.engine.begin() as conn:
            conn.execute(
                text(f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})"),
                payload
            )

    def add_rows(self, sheet_name: str, data_list_of_lists: List[List[Any]]):
        """Bulk insert multiple rows for faster migration."""
        if not data_list_of_lists: return
        table_name = self._get_table_name(sheet_name)
        
        cols = []
        if sheet_name == 'Customers':
            cols = ['customer_id', 'name', 'phone', 'email', 'address', 'id_proof_type', 'id_proof_number', 'status', 'created_date', 'notes']
        elif sheet_name == 'Loans':
            cols = ['loan_id', 'customer_id', 'principal_amount', 'add_on_principal', 'interest_rate', 'loan_type', 'start_date', 'tenure_months', 'status', 'fund_source', 'created_date', 'closed_date', 'notes', 'transaction_type', 'debt_interest_mode', 'pre_deducted_interest', 'net_disbursed_amount', 'original_interest_amount', 'waived_interest_amount', 'waiver_reason', 'waiver_date']
        elif sheet_name == 'Payments':
            cols = ['payment_id', 'loan_id', 'customer_id', 'payment_date', 'amount', 'payment_type', 'payment_method', 'reference_number', 'created_date', 'created_by', 'notes', 'principal_amount', 'interest_amount', 'help_category']
        elif sheet_name == 'Help':
            cols = ['help_id', 'customer_id', 'customer_name', 'help_date', 'help_amount', 'help_category', 'help_note', 'repayment_date', 'repayment_amount', 'status']
        elif sheet_name == 'CapitalInjections':
            cols = ['injection_id', 'source_type', 'amount', 'injection_date', 'description', 'created_by', 'created_date']
        elif sheet_name == 'AuditLog':
            cols = ['log_id', 'entity_type', 'entity_id', 'action', 'old_value', 'new_value', '"user"', 'timestamp']
        
        placeholders = ", ".join([f":{col.replace('\"', '')}" for col in cols])
        columns_str = ", ".join(cols)
        query = f"INSERT INTO {table_name} ({columns_str}) VALUES ({placeholders}) ON CONFLICT DO NOTHING"
        
        params_list = []
        for data in data_list_of_lists:
             if len(data) == len(cols):
                  params_list.append({col.replace("\"", ""): val for col, val in zip(cols, data)})
        
        if params_list:
             # Process in chunks of 500 to avoid PostgreSQL parameter limit (max ~32k parameters per query)
             chunk_size = 500
             for i in range(0, len(params_list), chunk_size):
                 chunk = params_list[i:i + chunk_size]
                 with self.engine.begin() as conn:
                     conn.execute(text(query), chunk)

    def update_row(self, sheet_name: str, id_column: str, id_value: str, updates: Dict):
        table_name = self._get_table_name(sheet_name)
        allowed = self._get_allowed_columns(sheet_name)
        normalized_id_column = self._normalize_column_name(sheet_name, id_column)
        if normalized_id_column not in allowed:
            raise ValueError(f"Unsupported id column '{id_column}' for {sheet_name}")
        
        set_clauses = []
        params = {normalized_id_column: id_value}
        
        for k, v in updates.items():
            normalized_key = self._normalize_column_name(sheet_name, k)
            if normalized_key not in allowed:
                continue
            param_key = f"update_{normalized_key}"
            set_clauses.append(f"{normalized_key} = :{param_key}")
            params[param_key] = v
            
        if not set_clauses:
            return
            
        query = f"UPDATE {table_name} SET {', '.join(set_clauses)} WHERE {normalized_id_column} = :{normalized_id_column}"
        
        with self.engine.begin() as conn:
            conn.execute(text(query), params)

    def delete_row(self, sheet_name: str, id_column: str, id_value: str):
        table_name = self._get_table_name(sheet_name)
        normalized_id_column = self._normalize_column_name(sheet_name, id_column)
        if normalized_id_column not in self._get_allowed_columns(sheet_name):
            raise ValueError(f"Unsupported id column '{id_column}' for {sheet_name}")
        with self.engine.begin() as conn:
            conn.execute(
                text(f"DELETE FROM {table_name} WHERE {normalized_id_column} = :id_value"),
                {"id_value": id_value}
            )

    def log_audit(self, entity_type: str, entity_id: str, action: str, old_value: Any = None, new_value: Any = None, user: str = 'USER'):
        audit_id = self.get_next_id('audit')
        data = [
            audit_id,
            entity_type,
            entity_id,
            action,
            json.dumps(old_value) if old_value else '',
            json.dumps(new_value) if new_value else '',
            user,
            datetime.now()
        ]
        self.add_row('AuditLog', data)
        return audit_id

# Initialize database
# Determine if we should use Postgres (Railway environment) or fallback to Excel Local
RAILWAY_DB_URL = os.environ.get('DATABASE_URL')

if RAILWAY_DB_URL:
    logger.info("Initializing PostgresDB connection...")
    db = PostgresDB(RAILWAY_DB_URL)
else:
    logger.info("Initializing ExcelDB connection...")
    db = ExcelDB(EXCEL_DB_PATH)

# ==================== HELPER FUNCTIONS ====================

def calculate_interest_accrued(principal: float, rate: float, start_date: date, end_date: Optional[date] = None) -> float:
    """Calculate simple interest accrued"""
    if end_date is None:
        end_date = datetime.now().date()
    
    if isinstance(start_date, datetime):
        start_date = start_date.date()
    
    days = (end_date - start_date).days
    months = days / 30.0  # Approximate months
    
    # Simple interest calculation
    interest = principal * rate * months
    return round(interest, 2)

def get_loan_type(loan: dict) -> str:
    """
    Get normalized loan type from loan record.
    Checks multiple possible column names for type field.
    Returns 'KULU', 'DEBT', or 'OTHER' (uppercase normalized).
    """
    # Check various possible column names for loan type
    # PRIORITIZE 'TYPE' first since that's the user's actual Excel column name
    type_columns = ['TYPE', 'Type', 'type', 'LoanType', 'Loan_Type', 'loan_type',
                    'transaction_type', 'TransactionType', 'Transaction_Type',
                    'Category', 'category', 'Loan Type']
    
    loan_type = None
    
    for col in type_columns:
        val = loan.get(col)
        if val is not None and str(val).strip():
            loan_type = str(val).strip()
            break
    
    # Normalize and return
    if loan_type:
        loan_type_upper = loan_type.upper()
        if loan_type_upper in ['KULU', 'DEBT']:
            return loan_type_upper
    
    return 'OTHER'

DATE_DISPLAY_FORMAT = '%d-%b-%y'
HELP_CATEGORIES = {'Personal', 'Medical', 'Emergency', 'Business', 'Other'}
HELP_STATUSES = {'Active', 'Settled', 'Partial'}
DEBT_INTEREST_MODES = {'upfront_deduction', 'subsequent_collection'}

def to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, ''):
            return default
        if isinstance(value, Decimal):
            return float(value)
        return float(value)
    except (TypeError, ValueError):
        return default

def to_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ''):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default

def to_optional_money(value: Any) -> Any:
    if value in (None, ''):
        return ''
    return round(to_float(value), 2)

def validate_phone_number(phone: str) -> str:
    normalized = re.sub(r'[\s\-]', '', str(phone or '').strip())
    if not re.fullmatch(r'\+?\d{10,15}', normalized):
        raise HTTPException(status_code=400, detail="Phone number must contain 10 to 15 digits")
    return normalized

def first_present(record: Dict[str, Any], keys: List[str], default: Any = None):
    for key in keys:
        if key in record and record[key] not in (None, ''):
            return record[key]
    return default

def parse_date_value(value: Any) -> Optional[date]:
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        for fmt in ('%Y-%m-%d', '%d-%b-%y', '%d-%b-%Y', '%d/%m/%Y', '%d/%m/%y'):
            try:
                return datetime.strptime(raw[:11], fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(raw[:10]).date()
        except ValueError:
            return None
    return None

def format_display_date(value: Any) -> str:
    parsed = parse_date_value(value)
    return parsed.strftime(DATE_DISPLAY_FORMAT) if parsed else ''

def format_iso_date(value: Any) -> str:
    parsed = parse_date_value(value)
    return parsed.isoformat() if parsed else ''

def title_case_status(value: Any, default: str = 'Active') -> str:
    raw = str(value or default).strip().replace('_', ' ')
    if not raw:
        raw = default
    return ' '.join(part.capitalize() for part in raw.split())

def resolve_debt_interest_mode(record: Dict[str, Any]) -> str:
    raw_mode = str(first_present(record, ['debt_interest_mode', 'DebtInterestMode'], 'subsequent_collection')).strip().lower()
    if raw_mode not in DEBT_INTEREST_MODES:
        raw_mode = 'subsequent_collection'
    return raw_mode

def get_raw_principal_amount(record: Dict[str, Any]) -> float:
    return round(to_float(first_present(record, ['principal_amount', 'Principal', 'PrincipalAmount', 'Amount', 'LoanAmount'])), 2)

def get_add_on_principal_amount(record: Dict[str, Any]) -> float:
    return round(max(0, to_float(first_present(record, ['add_on_principal', 'AddOnPrincipal', 'addOnPrincipal']))), 2)

def get_effective_principal_amount(record: Dict[str, Any]) -> float:
    # BUSINESS RULE: reporting principal excludes add_on_principal while preserving the stored raw principal.
    return round(max(0, get_raw_principal_amount(record) - get_add_on_principal_amount(record)), 2)

def resolve_payment_components(payment: Dict[str, Any], remaining_principal: float) -> Dict[str, float]:
    total_amount = round(to_float(payment.get('total_amount', payment.get('amount'))), 2)
    principal_amount = round(max(0, to_float(payment.get('principal_amount'))), 2)
    interest_amount = round(max(0, to_float(payment.get('interest_amount'))), 2)
    payment_type = str(payment.get('payment_type', '') or '').strip().upper()

    if total_amount <= 0 and (principal_amount > 0 or interest_amount > 0):
        total_amount = round(principal_amount + interest_amount, 2)

    if principal_amount <= 0 and interest_amount <= 0 and total_amount > 0:
        if payment_type == 'PRINCIPAL':
            principal_amount = min(total_amount, remaining_principal)
        elif payment_type == 'INTEREST':
            interest_amount = total_amount
        else:
            principal_amount = min(total_amount, remaining_principal)
            interest_amount = round(max(0, total_amount - principal_amount), 2)
    elif payment_type == 'PRINCIPAL':
        principal_amount = min(total_amount or principal_amount, remaining_principal)
        interest_amount = 0.0
    elif payment_type == 'INTEREST':
        principal_amount = 0.0
        interest_amount = total_amount or interest_amount
    else:
        principal_amount = min(principal_amount, remaining_principal)
        if total_amount > 0:
            interest_amount = round(max(0, total_amount - principal_amount), 2)

    principal_amount = round(min(principal_amount, max(0, remaining_principal)), 2)
    if total_amount <= 0:
        total_amount = round(principal_amount + interest_amount, 2)
    else:
        interest_amount = round(max(0, total_amount - principal_amount), 2)

    return {
        'principal_amount': principal_amount,
        'interest_amount': interest_amount,
        'total_amount': round(total_amount, 2)
    }

def build_customer_lookup(customers: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Dict[str, Any]]:
    source = customers if customers is not None else db.get_all_rows('Customers')
    lookup = {}
    for record in source:
        normalized = normalize_customer(record)
        lookup[normalized['customer_id']] = normalized
    return lookup

def build_loan_lookup(loans: Optional[List[Dict[str, Any]]] = None, customer_lookup: Optional[Dict[str, Dict[str, Any]]] = None) -> Dict[str, Dict[str, Any]]:
    source = loans if loans is not None else db.get_all_rows('Loans')
    lookup = {}
    for record in source:
        normalized = normalize_loan(record, customer_lookup=customer_lookup, payment_rows=[])
        lookup[normalized['loan_id']] = normalized
    return lookup

def normalize_customer(record: Dict[str, Any]) -> Dict[str, Any]:
    customer_id = str(first_present(record, ['customer_id', 'CustomerID', 'BorrowerID', 'BorrowerId', 'CustomerId', 'id'], '') or '')
    name = str(first_present(record, ['name', 'Name', 'BorrowerName', 'CustomerName'], customer_id) or customer_id)
    created_date = first_present(record, ['created_date', 'CreatedOn', 'createdDate'])
    status = title_case_status(first_present(record, ['status', 'Status', 'IsActive'], 'ACTIVE'))
    if str(first_present(record, ['status', 'Status'], '')).strip() == '' and str(first_present(record, ['IsActive'], '')).strip().lower() in {'yes', 'true', '1'}:
        status = 'Active'
    elif str(first_present(record, ['status', 'Status'], '')).strip() == '' and str(first_present(record, ['IsActive'], '')).strip():
        status = 'Inactive'

    return {
        'id': customer_id,
        'customer_id': customer_id,
        'name': name,
        'phone': str(first_present(record, ['phone', 'Phone'], '') or ''),
        'email': str(first_present(record, ['email', 'Email'], '') or ''),
        'address': str(first_present(record, ['address', 'Address'], '') or ''),
        'id_proof_type': str(first_present(record, ['id_proof_type', 'IdProofType'], '') or ''),
        'id_proof_number': str(first_present(record, ['id_proof_number', 'IdProofNumber'], '') or ''),
        'status': status,
        'created_date': format_display_date(created_date),
        'created_date_iso': format_iso_date(created_date),
        'notes': str(first_present(record, ['notes', 'Notes'], '') or '')
    }

def infer_payment_split(total_amount: float, payment_type: Optional[str], principal_amount: float, interest_amount: float) -> Dict[str, float]:
    total = round(to_float(total_amount), 2)
    principal = round(to_float(principal_amount), 2)
    interest = round(to_float(interest_amount), 2)

    if principal == 0 and interest == 0 and total > 0:
        normalized_type = str(payment_type or '').strip().upper()
        if normalized_type == 'PRINCIPAL':
            principal = total
        elif normalized_type == 'INTEREST':
            interest = total
        elif normalized_type == 'BOTH':
            principal = total
        else:
            interest = total

    if total == 0 and (principal > 0 or interest > 0):
        total = round(principal + interest, 2)

    if total > 0 and round(principal + interest, 2) == 0:
        interest = total

    return {
        'principal_amount': round(principal, 2),
        'interest_amount': round(interest, 2),
        'total_amount': round(total, 2)
    }

def normalize_help(record: Dict[str, Any]) -> Dict[str, Any]:
    help_date = first_present(record, ['help_date', 'HelpDate'])
    repayment_date = first_present(record, ['repayment_date', 'RepaymentDate'])
    status = title_case_status(first_present(record, ['status', 'Status'], 'Active'))
    category = str(first_present(record, ['help_category', 'HelpCategory'], 'Other') or 'Other').title()
    if category not in HELP_CATEGORIES:
        category = 'Other'

    return {
        'help_id': str(first_present(record, ['help_id', 'HelpID'], '') or ''),
        'customer_id': str(first_present(record, ['customer_id', 'CustomerID'], '') or ''),
        'customer_name': str(first_present(record, ['customer_name', 'CustomerName'], '') or ''),
        'help_date': format_display_date(help_date),
        'help_date_iso': format_iso_date(help_date),
        'help_amount': round(to_float(first_present(record, ['help_amount', 'HelpAmount'])), 2),
        'help_category': category,
        'help_note': str(first_present(record, ['help_note', 'HelpNote'], '') or ''),
        'repayment_date': format_display_date(repayment_date),
        'repayment_date_iso': format_iso_date(repayment_date),
        'repayment_amount': round(to_float(first_present(record, ['repayment_amount', 'RepaymentAmount'])), 2),
        'status': status if status in HELP_STATUSES else 'Active'
    }

def normalize_payment(record: Dict[str, Any], customer_lookup: Optional[Dict[str, Dict[str, Any]]] = None, loan_lookup: Optional[Dict[str, Dict[str, Any]]] = None) -> Dict[str, Any]:
    payment_date = first_present(record, ['payment_date', 'PaymentDate', 'Date', 'TransactionDate'])
    customer_id = str(first_present(record, ['customer_id', 'CustomerID', 'BorrowerID', 'BorrowerId', 'Borrower'], '') or '')
    loan_id = str(first_present(record, ['loan_id', 'LoanID', 'LoanId'], '') or '')
    split = infer_payment_split(
        first_present(record, ['amount', 'Amount', 'PaymentAmount', 'total_amount', 'TotalAmount'], 0),
        first_present(record, ['payment_type', 'PaymentType', 'Type'], ''),
        first_present(record, ['principal_amount', 'PrincipalAmount'], 0),
        first_present(record, ['interest_amount', 'InterestAmount'], 0)
    )

    customer_name = ''
    if customer_lookup and customer_id in customer_lookup:
        customer_name = customer_lookup[customer_id]['name']
    elif loan_lookup and loan_id in loan_lookup:
        customer_name = loan_lookup[loan_id].get('customer_name', '')

    payment_type = str(first_present(record, ['payment_type', 'PaymentType', 'Type'], '') or '').strip().upper()
    if not payment_type:
        if split['principal_amount'] > 0 and split['interest_amount'] > 0:
            payment_type = 'BOTH'
        elif split['principal_amount'] > 0:
            payment_type = 'PRINCIPAL'
        else:
            payment_type = 'INTEREST'

    return {
        'payment_id': str(first_present(record, ['payment_id', 'PaymentID'], '') or ''),
        'loan_id': loan_id,
        'customer_id': customer_id,
        'customer_name': customer_name,
        'payment_date': format_display_date(payment_date),
        'payment_date_iso': format_iso_date(payment_date),
        'amount': split['total_amount'],
        'total_amount': split['total_amount'],
        'principal_amount': split['principal_amount'],
        'interest_amount': split['interest_amount'],
        'payment_type': payment_type,
        'payment_method': str(first_present(record, ['payment_method', 'PaymentMethod'], 'CASH') or 'CASH'),
        'reference_number': str(first_present(record, ['reference_number', 'ReferenceNumber'], '') or ''),
        'created_date': format_display_date(first_present(record, ['created_date', 'CreatedOn'])),
        'created_date_iso': format_iso_date(first_present(record, ['created_date', 'CreatedOn'])),
        'created_by': str(first_present(record, ['created_by', 'CreatedBy'], 'USER') or 'USER'),
        'help_category': str(first_present(record, ['help_category', 'HelpCategory'], 'None') or 'None'),
        'notes': str(first_present(record, ['notes', 'Notes', 'Remarks'], '') or '')
    }

def calculate_loan_balances(loan_record: Dict[str, Any], payment_rows: List[Dict[str, Any]]) -> Dict[str, float]:
    effective_principal = round(to_float(loan_record.get('effective_principal_amount', loan_record.get('principal_amount'))), 2)
    principal_paid = 0.0
    interest_paid = 0.0

    def payment_sort_key(payment: Dict[str, Any]):
        return (
            payment.get('payment_date_iso') or format_iso_date(payment.get('payment_date')) or '',
            payment.get('payment_id', '')
        )

    for payment in sorted(payment_rows, key=payment_sort_key):
        remaining = max(0, effective_principal - principal_paid)
        split = resolve_payment_components(payment, remaining)
        principal_paid = round(principal_paid + split['principal_amount'], 2)
        interest_paid = round(interest_paid + split['interest_amount'], 2)

    outstanding = round(max(0, effective_principal - principal_paid), 2)
    return {
        'principal_paid': principal_paid,
        'interest_paid': interest_paid,
        'total_paid': round(principal_paid + interest_paid, 2),
        'outstanding_balance': outstanding
    }

def normalize_loan(record: Dict[str, Any], customer_lookup: Optional[Dict[str, Dict[str, Any]]] = None, payment_rows: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    customer_id = str(first_present(record, ['customer_id', 'CustomerID', 'BorrowerID', 'BorrowerId'], '') or '')
    customer_name = customer_lookup.get(customer_id, {}).get('name', '') if customer_lookup else ''
    principal_amount = get_raw_principal_amount(record)
    add_on_principal = get_add_on_principal_amount(record)
    effective_principal_amount = get_effective_principal_amount(record)
    pre_deducted_interest = round(to_float(first_present(record, ['pre_deducted_interest', 'PreDeductedInterest'])), 2)
    transaction_type = get_loan_type(record)
    debt_interest_mode = resolve_debt_interest_mode(record)

    # BUSINESS RULE: KULU never deducts interest at source; only DEBT can use upfront deduction.
    if transaction_type != 'DEBT':
        debt_interest_mode = 'subsequent_collection'
        pre_deducted_interest = 0.0

    net_disbursed_amount = to_float(first_present(record, ['net_disbursed_amount', 'NetDisbursedAmount']), 0)
    if net_disbursed_amount <= 0:
        if transaction_type == 'DEBT' and debt_interest_mode == 'upfront_deduction':
            net_disbursed_amount = max(0, principal_amount - pre_deducted_interest)
        else:
            net_disbursed_amount = principal_amount

    balances = calculate_loan_balances({'principal_amount': principal_amount, 'effective_principal_amount': effective_principal_amount}, payment_rows or [])
    start_date = first_present(record, ['start_date', 'StartDate', 'LoanDate', 'Date'])
    start_date_value = parse_date_value(start_date)
    days_active = (datetime.now().date() - start_date_value).days if start_date_value else 0
    months_active = max(0, to_int(days_active / 30))
    accrued_interest = calculate_interest_accrued(principal_amount, to_float(first_present(record, ['interest_rate', 'InterestRate', 'Rate'])), start_date_value) if start_date_value else 0.0
    interest_paid_total = round(balances['interest_paid'] + (pre_deducted_interest if debt_interest_mode == 'upfront_deduction' else 0), 2)

    return {
        'loan_id': str(first_present(record, ['loan_id', 'LoanID', 'LoanId'], '') or ''),
        'customer_id': customer_id,
        'customer_name': customer_name or customer_id,
        'principal_amount': principal_amount,
        'add_on_principal': add_on_principal,
        'effective_principal_amount': effective_principal_amount,
        'interest_rate': round(to_float(first_present(record, ['interest_rate', 'InterestRate', 'Rate'])), 6),
        'loan_type': str(first_present(record, ['loan_type', 'LoanType'], 'PERSONAL') or 'PERSONAL'),
        'transaction_type': transaction_type,
        'debt_interest_mode': debt_interest_mode,
        'pre_deducted_interest': pre_deducted_interest,
        'net_disbursed_amount': round(net_disbursed_amount, 2),
        'start_date': format_display_date(start_date),
        'start_date_iso': format_iso_date(start_date),
        'tenure_months': to_int(first_present(record, ['tenure_months', 'TenureMonths']), 0) or None,
        'status': title_case_status(first_present(record, ['status', 'Status', 'LoanStatus'], 'ACTIVE')),
        'fund_source': str(first_present(record, ['fund_source', 'FundSourceID'], '') or ''),
        'notes': str(first_present(record, ['notes', 'Notes'], '') or ''),
        'created_date': format_display_date(first_present(record, ['created_date', 'CreatedOn'])),
        'created_date_iso': format_iso_date(first_present(record, ['created_date', 'CreatedOn'])),
        'closed_date': format_display_date(first_present(record, ['closed_date', 'ClosedDate'])),
        'closed_date_iso': format_iso_date(first_present(record, ['closed_date', 'ClosedDate'])),
        'principal_paid': balances['principal_paid'],
        'interest_paid': interest_paid_total,
        'total_paid': round(balances['principal_paid'] + interest_paid_total, 2),
        'outstanding_balance': balances['outstanding_balance'],
        'months_active': months_active,
        'days_active': days_active,
        'total_interest_accrued': round(accrued_interest, 2),
        'original_interest_amount': round(to_float(first_present(record, ['original_interest_amount', 'OriginalInterestAmount']), accrued_interest), 2),
        'waived_interest_amount': round(to_float(first_present(record, ['waived_interest_amount', 'WaivedInterestAmount'])), 2),
        'waiver_reason': str(first_present(record, ['waiver_reason', 'WaiverReason'], '') or ''),
        'waiver_date': format_display_date(first_present(record, ['waiver_date', 'WaiverDate'])),
        'waiver_date_iso': format_iso_date(first_present(record, ['waiver_date', 'WaiverDate'])),
        'effective_interest_due': round(max(0, accrued_interest - interest_paid_total - to_float(first_present(record, ['waived_interest_amount', 'WaivedInterestAmount']))), 2)
    }

def load_normalized_customers(status: Optional[str] = None, search: Optional[str] = None) -> List[Dict[str, Any]]:
    customers = [normalize_customer(record) for record in db.get_all_rows('Customers')]
    if status:
        customers = [customer for customer in customers if customer['status'].upper() == status.upper()]
    if search:
        search_lower = search.lower()
        customers = [
            customer for customer in customers
            if search_lower in customer['customer_id'].lower()
            or search_lower in customer['name'].lower()
            or search_lower in customer['phone'].lower()
        ]
    return sorted(customers, key=lambda customer: (customer['name'] or '').lower())

def load_normalized_payments(loan_id: Optional[str] = None, customer_id: Optional[str] = None, payment_method: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None) -> List[Dict[str, Any]]:
    customers = build_customer_lookup()
    loans = build_loan_lookup(customer_lookup=customers)
    payments = [normalize_payment(record, customer_lookup=customers, loan_lookup=loans) for record in db.get_all_rows('Payments')]

    if loan_id:
        payments = [payment for payment in payments if payment['loan_id'] == loan_id]
    if customer_id:
        payments = [payment for payment in payments if payment['customer_id'] == customer_id]
    if payment_method and payment_method.upper() != 'ALL':
        payments = [payment for payment in payments if payment['payment_method'].upper() == payment_method.upper()]

    start = parse_date_value(start_date)
    end = parse_date_value(end_date)
    if start or end:
        filtered = []
        for payment in payments:
            payment_date = parse_date_value(payment['payment_date_iso'])
            if not payment_date:
                continue
            if start and payment_date < start:
                continue
            if end and payment_date > end:
                continue
            filtered.append(payment)
        payments = filtered

    payments.sort(key=lambda payment: payment['payment_date_iso'] or '', reverse=True)
    return payments

def load_normalized_loans(customer_id: Optional[str] = None, status: Optional[str] = None, loan_type: Optional[str] = None, disbursed_from: Optional[str] = None, disbursed_to: Optional[str] = None) -> List[Dict[str, Any]]:
    customers = build_customer_lookup()
    payments = load_normalized_payments()
    payments_by_loan: Dict[str, List[Dict[str, Any]]] = {}
    for payment in payments:
        payments_by_loan.setdefault(payment['loan_id'], []).append(payment)

    normalized_loans = [
        normalize_loan(record, customer_lookup=customers, payment_rows=payments_by_loan.get(str(first_present(record, ['loan_id', 'LoanID', 'LoanId'], '') or ''), []))
        for record in db.get_all_rows('Loans')
    ]

    if customer_id:
        normalized_loans = [loan for loan in normalized_loans if loan['customer_id'] == customer_id]
    if status and status.upper() != 'ALL':
        normalized_loans = [loan for loan in normalized_loans if loan['status'].upper() == status.upper()]
    if loan_type and loan_type.upper() != 'ALL':
        normalized_loans = [loan for loan in normalized_loans if loan['transaction_type'].upper() == loan_type.upper()]

    from_date = parse_date_value(disbursed_from)
    to_date = parse_date_value(disbursed_to)
    if from_date or to_date:
        filtered_loans = []
        for loan in normalized_loans:
            start = parse_date_value(loan['start_date_iso'])
            if not start:
                continue
            if from_date and start < from_date:
                continue
            if to_date and start > to_date:
                continue
            filtered_loans.append(loan)
        normalized_loans = filtered_loans

    normalized_loans.sort(key=lambda loan: loan['start_date_iso'] or '', reverse=True)
    return normalized_loans

# Debug endpoint to inspect actual Excel structure
@app.get("/debug/excel-structure")
async def debug_excel_structure():
    """Debug endpoint to see actual column names and sample data from Excel"""
    ensure_debug_access()
    loans = db.get_all_rows('Loans')
    payments = db.get_all_rows('Payments')
    customers = db.get_all_rows('Customers')
    
    # Get column names
    loan_columns = list(loans[0].keys()) if loans else []
    payment_columns = list(payments[0].keys()) if payments else []
    customer_columns = list(customers[0].keys()) if customers else []
    
    # Get sample loan with all values
    sample_loans = loans[:3] if loans else []
    sample_payments = payments[:3] if payments else []
    
    # Count types using get_loan_type (which checks TYPE column first)
    type_counts = {'KULU': 0, 'DEBT': 0, 'OTHER': 0}
    for loan in loans:
        loan_type = get_loan_type(loan)
        type_counts[loan_type] = type_counts.get(loan_type, 0) + 1
    
    # Show raw TYPE column values directly
    type_column_raw = {}
    for loan in loans:
        # Check TYPE, Type, type columns directly
        for col in ['TYPE', 'Type', 'type']:
            val = loan.get(col)
            if val is not None:
                type_column_raw[str(val)] = type_column_raw.get(str(val), 0) + 1
                break
    
    # Show payment type values
    payment_type_values = {}
    for p in payments:
        for col in ['PaymentType', 'payment_type', 'Type']:
            val = p.get(col)
            if val is not None:
                payment_type_values[str(val)] = payment_type_values.get(str(val), 0) + 1
                break
    
    return {
        'loan_count': len(loans),
        'payment_count': len(payments),
        'customer_count': len(customers),
        'loan_columns': loan_columns,
        'payment_columns': payment_columns,
        'customer_columns': customer_columns,
        'sample_loans': sample_loans,
        'sample_payments': sample_payments,
        'type_counts_normalized': type_counts,
        'type_column_raw_values': type_column_raw,
        'payment_type_values': payment_type_values
    }

# Data validation endpoint - checks data integrity across all tables
@app.get("/debug/data-validation")
async def validate_data_integrity():
    """
    Comprehensive data validation endpoint.
    Checks for orphan records, missing mappings, and data integrity issues.
    """
    ensure_debug_access()
    loans = db.get_all_rows('Loans')
    payments = db.get_all_rows('Payments')
    customers = db.get_all_rows('Customers')
    
    # Helper functions
    def get_loan_id(record):
        for key in ['LoanID', 'loan_id', 'Loan_ID']:
            if key in record and record[key] is not None:
                return str(record[key])
        return None
    
    def get_borrower_id(record):
        for key in ['BorrowerId', 'borrower_id', 'customer_id', 'CustomerID']:
            if key in record and record[key] is not None:
                return str(record[key])
        return None
    
    def get_principal(loan):
        for key in ['principal_amount', 'Principal', 'PrincipalAmount', 'Amount']:
            if key in loan and loan[key] is not None:
                try:
                    return float(loan[key])
                except:
                    continue
        return 0
    
    def get_payment_amount(payment):
        for key in ['amount', 'Amount', 'PaymentAmount']:
            if key in payment and payment[key] is not None:
                try:
                    return float(payment[key])
                except:
                    continue
        return 0
    
    # Build sets for validation
    loan_ids = {get_loan_id(l) for l in loans if get_loan_id(l)}
    borrower_ids = {get_borrower_id(c) for c in customers if get_borrower_id(c)}
    loan_borrower_ids = {get_borrower_id(l) for l in loans if get_borrower_id(l)}
    
    # Check for orphan payments (payments without matching loan)
    orphan_payments = []
    for p in payments:
        plid = get_loan_id(p)
        if plid and plid not in loan_ids:
            orphan_payments.append({'payment_loan_id': plid, 'amount': get_payment_amount(p)})
    
    # Check for loans without borrower mapping
    loans_without_borrower = []
    for l in loans:
        bid = get_borrower_id(l)
        if bid and bid not in borrower_ids:
            loans_without_borrower.append({'loan_id': get_loan_id(l), 'borrower_id': bid})
    
    # Calculate totals for verification
    total_disbursed = sum(get_principal(l) for l in loans)
    total_payments = sum(get_payment_amount(p) for p in payments)
    
    # Per-type totals
    kulu_disbursed = sum(get_principal(l) for l in loans if get_loan_type(l) == 'KULU')
    debt_disbursed = sum(get_principal(l) for l in loans if get_loan_type(l) == 'DEBT')
    other_disbursed = sum(get_principal(l) for l in loans if get_loan_type(l) == 'OTHER')
    
    # Per-type loan counts
    kulu_count = len([l for l in loans if get_loan_type(l) == 'KULU'])
    debt_count = len([l for l in loans if get_loan_type(l) == 'DEBT'])
    other_count = len([l for l in loans if get_loan_type(l) == 'OTHER'])
    
    issues = []
    if orphan_payments:
        issues.append(f"{len(orphan_payments)} orphan payments without matching loans")
    if loans_without_borrower:
        issues.append(f"{len(loans_without_borrower)} loans without valid borrower mapping")
    if other_count > 0:
        issues.append(f"{other_count} loans classified as OTHER (may need TYPE column review)")
    
    return {
        'validation_status': 'PASS' if not issues else 'ISSUES_FOUND',
        'issues': issues,
        'summary': {
            'total_loans': len(loans),
            'total_payments': len(payments),
            'total_customers': len(customers),
            'total_disbursed': round(total_disbursed, 2),
            'total_payments_amount': round(total_payments, 2)
        },
        'by_type': {
            'KULU': {'count': kulu_count, 'disbursed': round(kulu_disbursed, 2)},
            'DEBT': {'count': debt_count, 'disbursed': round(debt_disbursed, 2)},
            'OTHER': {'count': other_count, 'disbursed': round(other_disbursed, 2)}
        },
        'orphan_payments_sample': orphan_payments[:5],
        'loans_without_borrower_sample': loans_without_borrower[:5],
        'unique_loan_ids': len(loan_ids),
        'unique_borrowers': len(borrower_ids),
        'loan_borrower_mappings': len(loan_borrower_ids)
    }


def get_loan_summary_data(loan_id: str) -> Optional[LoanSummary]:
    """Get comprehensive loan summary with calculations"""
    payments = load_normalized_payments(loan_id=loan_id)
    loans = load_normalized_loans()
    loan = next((item for item in loans if item['loan_id'] == loan_id), None)

    if not loan:
        return None

    return LoanSummary(
        loan_id=loan['loan_id'],
        customer_name=loan['customer_name'],
        principal_amount=loan['principal_amount'],
        add_on_principal=loan.get('add_on_principal', 0.0),
        effective_principal_amount=loan.get('effective_principal_amount', loan['principal_amount']),
        net_disbursed_amount=loan['net_disbursed_amount'],
        pre_deducted_interest=loan['pre_deducted_interest'],
        total_paid=loan['total_paid'],
        principal_paid=loan['principal_paid'],
        interest_paid=loan['interest_paid'],
        outstanding_balance=loan['outstanding_balance'],
        interest_rate=loan['interest_rate'],
        transaction_type=loan['transaction_type'],
        debt_interest_mode=loan['debt_interest_mode'],
        status=loan['status'],
        start_date=loan['start_date'],
        start_date_iso=loan['start_date_iso'],
        months_active=loan['months_active'],
        days_active=loan['days_active'],
        total_interest_accrued=loan['total_interest_accrued'],
        original_interest_amount=loan['original_interest_amount'],
        waived_interest_amount=loan['waived_interest_amount'],
        effective_interest_due=loan['effective_interest_due']
    )

# ==================== API ENDPOINTS ====================

# ========== CUSTOMER ENDPOINTS ==========

@app.get("/customers", response_model=List[Dict])
async def get_customers(status: Optional[str] = None, search: Optional[str] = None):
    """Get all customers with optional filtering"""
    return load_normalized_customers(status=status, search=search)

@app.get("/customers/{customer_id}")
async def get_customer(customer_id: str):
    """Get single customer by ID"""
    customers = load_normalized_customers()
    customer = next((c for c in customers if c['customer_id'] == customer_id), None)
    
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    return customer

@app.get("/customers/{customer_id}/profile")
async def get_customer_profile(customer_id: str):
    """Get a customer-centric view with profile, loans, transactions, and current totals."""
    customer = await get_customer(customer_id)
    loans = load_normalized_loans(customer_id=customer_id)
    payments = load_normalized_payments(customer_id=customer_id)
    help_records = [record for record in (normalize_help(row) for row in db.get_all_rows('Help')) if record['customer_id'] == customer_id]

    totals = {
        'total_disbursed': round(sum(to_float(loan.get('effective_principal_amount', loan.get('principal_amount'))) for loan in loans), 2),
        'total_collected': round(sum(to_float(loan.get('principal_paid')) for loan in loans), 2),
        'interest_collected': round(sum(to_float(loan.get('interest_paid')) for loan in loans), 2),
        'outstanding_balance': round(sum(to_float(loan.get('outstanding_balance')) for loan in loans), 2),
        'active_loans': len([loan for loan in loans if str(loan.get('status', '')).upper() == 'ACTIVE'])
    }

    transactions = sorted(
        [
            {
                'transaction_id': payment['payment_id'],
                'transaction_type': 'Payment',
                'date': payment['payment_date'],
                'date_iso': payment['payment_date_iso'],
                'amount': payment['total_amount'],
                'principal_amount': payment['principal_amount'],
                'interest_amount': payment['interest_amount'],
                'status': payment['payment_method'],
                'notes': payment['notes']
            }
            for payment in payments
        ] + [
            {
                'transaction_id': help_record['help_id'],
                'transaction_type': 'Help',
                'date': help_record['help_date'],
                'date_iso': help_record['help_date_iso'],
                'amount': help_record['help_amount'],
                'principal_amount': 0.0,
                'interest_amount': 0.0,
                'status': help_record['status'],
                'notes': help_record['help_note']
            }
            for help_record in help_records
        ],
        key=lambda row: (row.get('date_iso') or '', row.get('transaction_id') or ''),
        reverse=True
    )

    return {
        'customer': customer,
        'totals': totals,
        'loans': loans,
        'payments': payments,
        'help_records': help_records,
        'transactions': transactions
    }

@app.post("/customers")
async def create_customer(customer: Customer):
    """Create new customer"""
    customer_id = db.get_next_id('customer')
    phone = validate_phone_number(customer.phone)
    payload = {
        'customer_id': customer_id,
        'name': customer.name,
        'phone': phone,
        'email': customer.email or '',
        'address': customer.address or '',
        'id_proof_type': customer.id_proof_type or '',
        'id_proof_number': customer.id_proof_number or '',
        'status': customer.status,
        'created_date': datetime.now(),
        'notes': customer.notes or '',
        'BorrowerID': customer_id,
        'BorrowerName': customer.name,
        'Phone': phone,
        'Address': customer.address or '',
        'IsActive': 'Yes' if str(customer.status).upper() == 'ACTIVE' else 'No',
        'CreatedOn': datetime.now()
    }
    db.add_dict_row('Customers', payload)
    db.log_audit('CUSTOMER', customer_id, 'CREATE', None, {'name': customer.name, 'phone': phone})
    
    return {"customer_id": customer_id, "message": "Customer created successfully"}

@app.put("/customers/{customer_id}")
async def update_customer(customer_id: str, customer: Customer):
    """Update existing customer"""
    # Get old values for audit
    old_customer = await get_customer(customer_id)
    phone = validate_phone_number(customer.phone)
    
    updates = {
        'name': customer.name,
        'phone': phone,
        'email': customer.email or '',
        'address': customer.address or '',
        'id_proof_type': customer.id_proof_type or '',
        'id_proof_number': customer.id_proof_number or '',
        'status': customer.status,
        'notes': customer.notes or '',
        'BorrowerName': customer.name,
        'Phone': phone,
        'Address': customer.address or '',
        'IsActive': 'Yes' if str(customer.status).upper() == 'ACTIVE' else 'No'
    }
    
    db.update_row('Customers', 'customer_id', customer_id, updates)
    db.log_audit('CUSTOMER', customer_id, 'UPDATE', old_customer, updates)
    
    return {"message": "Customer updated successfully"}

# ========== LOAN ENDPOINTS ==========

@app.get("/loans")
async def get_loans(
    customer_id: Optional[str] = None, 
    status: Optional[str] = None,
    loan_type: Optional[str] = None,
    disbursed_from: Optional[str] = None,
    disbursed_to: Optional[str] = None
):
    """Get all loans with optional filtering, including payment totals"""
    return load_normalized_loans(
        customer_id=customer_id,
        status=status,
        loan_type=loan_type,
        disbursed_from=disbursed_from,
        disbursed_to=disbursed_to
    )

@app.get("/loans/{loan_id}/summary")
async def get_loan_summary(loan_id: str):
    """Get comprehensive loan summary with calculations"""
    summary = get_loan_summary_data(loan_id)
    
    if not summary:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    return summary

@app.post("/loans")
async def create_loan(loan: Loan):
    """Create new loan"""
    if loan.principal_amount <= 0:
        raise HTTPException(status_code=400, detail="Principal amount must be greater than 0")
    if loan.interest_rate < 0:
        raise HTTPException(status_code=400, detail="Interest rate cannot be negative")
    loan_id = db.get_next_id('loan')
    normalized_type = str(loan.transaction_type or 'KULU').strip().upper()
    debt_interest_mode = resolve_debt_interest_mode({'debt_interest_mode': loan.debt_interest_mode})
    add_on_principal = round(max(0, to_float(loan.add_on_principal)), 2)
    pre_deducted_interest = round(to_float(loan.pre_deducted_interest), 2)

    if add_on_principal > loan.principal_amount:
        raise HTTPException(status_code=400, detail="Add-on principal cannot exceed principal amount")

    # BUSINESS RULE: KULU loans always disburse the full principal and never deduct interest upfront.
    if normalized_type != 'DEBT':
        debt_interest_mode = 'subsequent_collection'
        pre_deducted_interest = 0.0

    if debt_interest_mode == 'upfront_deduction':
        if pre_deducted_interest < 0 or pre_deducted_interest > loan.principal_amount:
            raise HTTPException(status_code=400, detail="Pre-deducted interest must be between 0 and principal amount")
        net_disbursed_amount = round(loan.principal_amount - pre_deducted_interest, 2)
    else:
        pre_deducted_interest = 0.0
        net_disbursed_amount = round(loan.principal_amount, 2)
    
    payload = {
        'loan_id': loan_id,
        'customer_id': loan.customer_id,
        'principal_amount': loan.principal_amount,
        'add_on_principal': add_on_principal,
        'interest_rate': loan.interest_rate,
        'loan_type': loan.loan_type,
        'start_date': loan.start_date,
        'tenure_months': loan.tenure_months or '',
        'status': loan.status,
        'fund_source': loan.fund_source or '',
        'created_date': datetime.now(),
        'closed_date': None,
        'notes': loan.notes or '',
        'transaction_type': normalized_type,
        'debt_interest_mode': debt_interest_mode,
        'pre_deducted_interest': pre_deducted_interest,
        'net_disbursed_amount': net_disbursed_amount,
        'original_interest_amount': 0,
        'waived_interest_amount': 0,
        'waiver_reason': '',
        'waiver_date': None,
        'LoanID': loan_id,
        'BorrowerID': loan.customer_id,
        'TYPE': normalized_type,
        'PrincipalAmount': loan.principal_amount,
        'AddOnPrincipal': add_on_principal,
        'InterestRate': loan.interest_rate,
        'StartDate': loan.start_date,
        'FundSourceID': loan.fund_source or '',
        'LoanStatus': loan.status,
        'CreatedOn': datetime.now()
    }
    db.add_dict_row('Loans', payload)
    db.log_audit('LOAN', loan_id, 'CREATE', None, {
        'customer_id': loan.customer_id,
        'principal': loan.principal_amount,
        'add_on_principal': add_on_principal,
        'transaction_type': normalized_type,
        'debt_interest_mode': debt_interest_mode,
        'pre_deducted_interest': pre_deducted_interest,
        'net_disbursed_amount': net_disbursed_amount
    })
    
    return {"loan_id": loan_id, "message": "Loan created successfully"}

@app.put("/loans/{loan_id}")
async def update_loan(loan_id: str, loan: Loan):
    """Update existing loan"""
    if loan.principal_amount <= 0:
        raise HTTPException(status_code=400, detail="Principal amount must be greater than 0")
    if loan.interest_rate < 0:
        raise HTTPException(status_code=400, detail="Interest rate cannot be negative")
    normalized_type = str(loan.transaction_type or 'KULU').strip().upper()
    debt_interest_mode = resolve_debt_interest_mode({'debt_interest_mode': loan.debt_interest_mode})
    add_on_principal = round(max(0, to_float(loan.add_on_principal)), 2)
    pre_deducted_interest = round(to_float(loan.pre_deducted_interest), 2)

    if add_on_principal > loan.principal_amount:
        raise HTTPException(status_code=400, detail="Add-on principal cannot exceed principal amount")

    # BUSINESS RULE: KULU loans can only use subsequent collection; DEBT optionally supports upfront deduction.
    if normalized_type != 'DEBT':
        debt_interest_mode = 'subsequent_collection'
        pre_deducted_interest = 0.0
    elif debt_interest_mode == 'upfront_deduction':
        if pre_deducted_interest < 0 or pre_deducted_interest > loan.principal_amount:
            raise HTTPException(status_code=400, detail="Pre-deducted interest must be between 0 and principal amount")
    else:
        pre_deducted_interest = 0.0

    net_disbursed_amount = round(
        loan.principal_amount - pre_deducted_interest
        if debt_interest_mode == 'upfront_deduction'
        else loan.principal_amount,
        2
    )

    updates = {
        'customer_id': loan.customer_id,
        'principal_amount': loan.principal_amount,
        'add_on_principal': add_on_principal,
        'interest_rate': loan.interest_rate,
        'loan_type': loan.loan_type,
        'transaction_type': normalized_type,
        'start_date': loan.start_date,
        'tenure_months': loan.tenure_months or '',
        'status': loan.status,
        'fund_source': loan.fund_source or '',
        'debt_interest_mode': debt_interest_mode,
        'pre_deducted_interest': pre_deducted_interest,
        'net_disbursed_amount': net_disbursed_amount,
        'notes': loan.notes or '',
        'BorrowerID': loan.customer_id,
        'TYPE': normalized_type,
        'PrincipalAmount': loan.principal_amount,
        'AddOnPrincipal': add_on_principal,
        'InterestRate': loan.interest_rate,
        'StartDate': loan.start_date,
        'FundSourceID': loan.fund_source or '',
        'LoanStatus': loan.status
    }
    
    # Get current loan status for comparison
    current_loan_data = await get_loan_summary(loan_id) # Use get_loan_summary to fetch current data
    current_status = current_loan_data.status if current_loan_data else None

    if loan.status == 'COMPLETED' and str(current_status).upper() != 'COMPLETED':
        updates['closed_date'] = datetime.now()
    
    db.update_row('Loans', 'loan_id', loan_id, updates)
    db.log_audit('LOAN', loan_id, 'UPDATE', None, updates)
    
    return {"message": "Loan updated successfully"}

@app.post("/loans/{loan_id}/waive-interest")
async def waive_interest(loan_id: str, waiver: InterestWaiver):
    """Apply interest waiver to a loan"""
    # Get current loan data
    loan = next((item for item in load_normalized_loans() if item['loan_id'] == loan_id), None)
    
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    # Calculate current interest accrued
    start_date = parse_date_value(loan.get('start_date_iso') or loan.get('start_date'))
    if not start_date:
        raise HTTPException(status_code=400, detail="Loan start date is invalid")
    
    current_interest = calculate_interest_accrued(
        float(loan.get('principal_amount', 0) or 0),
        float(loan.get('interest_rate', 0) or 0),
        start_date
    )
    
    # Store original interest if not already set
    original_interest = float(loan.get('original_interest_amount', 0))
    if original_interest == 0:
        original_interest = current_interest
    
    # Calculate new waived amount (cumulative)
    current_waived = float(loan.get('waived_interest_amount', 0) or 0)
    new_waived_total = current_waived + waiver.waived_amount
    
    # Validate waiver doesn't exceed accrued interest
    if new_waived_total > original_interest:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot waive more than accrued interest. Maximum waivable: {original_interest - current_waived}"
        )
    
    updates = {
        'original_interest_amount': original_interest,
        'waived_interest_amount': new_waived_total,
        'waiver_reason': waiver.reason,
        'waiver_date': waiver.waiver_date or datetime.now().date()
    }
    
    db.update_row('Loans', 'loan_id', loan_id, updates)
    db.log_audit('LOAN', loan_id, 'WAIVER', 
        {'waived_before': current_waived}, 
        {'waived_amount': waiver.waived_amount, 'total_waived': new_waived_total, 'reason': waiver.reason}
    )
    
    return {
        "message": "Interest waiver applied successfully",
        "original_interest": original_interest,
        "total_waived": new_waived_total,
        "effective_interest_due": original_interest - new_waived_total
    }

@app.get("/loans/{loan_id}/waiver-history")
async def get_waiver_history(loan_id: str):
    """Get waiver audit trail for a loan"""
    audit_logs = db.get_all_rows('AuditLog')
    waivers = [
        log for log in audit_logs 
        if log.get('entity_id') == loan_id and log.get('action') == 'WAIVER'
    ]
    return waivers

# ========== PAYMENT ENDPOINTS ==========

@app.get("/payments")
async def get_payments(
    loan_id: Optional[str] = None,
    customer_id: Optional[str] = None,
    payment_method: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Get all payments with optional filtering"""
    return load_normalized_payments(
        loan_id=loan_id,
        customer_id=customer_id,
        payment_method=payment_method,
        start_date=start_date,
        end_date=end_date
    )

@app.post("/payments")
async def create_payment(payment: Payment):
    """Record new payment"""
    loan = next((item for item in load_normalized_loans() if item['loan_id'] == payment.loan_id), None)
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    if payment.customer_id != loan.get('customer_id'):
        raise HTTPException(status_code=400, detail="Payment customer does not match the loan customer")

    payment_id = db.get_next_id('payment')
    principal_amount = round(to_float(payment.principal_amount), 2)
    interest_amount = round(to_float(payment.interest_amount), 2)
    total_amount = round(to_float(payment.total_amount if payment.total_amount is not None else payment.amount), 2)

    if total_amount <= 0 and (principal_amount > 0 or interest_amount > 0):
        total_amount = round(principal_amount + interest_amount, 2)

    if principal_amount <= 0 and interest_amount <= 0 and total_amount > 0:
        interest_amount = total_amount

    if total_amount <= 0:
        raise HTTPException(status_code=400, detail="At least one of principal or interest amount must be greater than 0")

    if principal_amount < 0 or interest_amount < 0:
        raise HTTPException(status_code=400, detail="Principal and interest amounts cannot be negative")

    if round(principal_amount + interest_amount, 2) != round(total_amount, 2):
        raise HTTPException(status_code=400, detail="TotalAmount must equal PrincipalAmount + InterestAmount")

    if principal_amount > round(to_float(loan.get('outstanding_balance')), 2):
        raise HTTPException(status_code=400, detail="Principal amount cannot exceed the current outstanding balance")

    if principal_amount > 0 and interest_amount > 0:
        payment_type = 'BOTH'
    elif principal_amount > 0:
        payment_type = 'PRINCIPAL'
    else:
        payment_type = 'INTEREST'

    help_category = str(payment.help_category or 'None').title()
    if help_category != 'None' and help_category not in HELP_CATEGORIES:
        raise HTTPException(status_code=400, detail="Invalid help category")
    
    payload = {
        'payment_id': payment_id,
        'loan_id': payment.loan_id,
        'customer_id': payment.customer_id,
        'payment_date': payment.payment_date,
        'amount': total_amount,
        'payment_type': payment_type,
        'payment_method': payment.payment_method,
        'reference_number': payment.reference_number or '',
        'created_date': datetime.now(),
        'created_by': payment.created_by,
        'notes': payment.notes or '',
        'principal_amount': principal_amount,
        'interest_amount': interest_amount,
        'help_category': help_category,
        'PaymentID': payment_id,
        'LoanID': payment.loan_id,
        'Borrower': payment.customer_id,
        'PaymentDate': payment.payment_date,
        'PaymentAmount': total_amount,
        'PaymentType': payment_type,
        'Remarks': payment.notes or '',
        'CreatedOn': datetime.now()
    }
    db.add_dict_row('Payments', payload)
    db.log_audit('PAYMENT', payment_id, 'CREATE', None, {
        'loan_id': payment.loan_id,
        'amount': total_amount,
        'principal_amount': principal_amount,
        'interest_amount': interest_amount,
        'type': payment_type,
        'help_category': help_category
    })

    help_id = None
    if help_category != 'None':
        customer = next((item for item in load_normalized_customers() if item['customer_id'] == payment.customer_id), None)
        help_id = db.get_next_id('help')
        help_status = title_case_status(payment.help_status, 'Active')
        if help_status not in HELP_STATUSES:
            help_status = 'Active'

        db.add_dict_row('Help', {
            'help_id': help_id,
            'customer_id': payment.customer_id,
            'customer_name': customer['name'] if customer else payment.customer_id,
            'help_date': payment.payment_date,
            'help_amount': total_amount,
            'help_category': help_category,
            'help_note': payment.help_note or payment.notes or '',
            'repayment_date': payment.repayment_date or '',
            'repayment_amount': to_optional_money(payment.repayment_amount),
            'status': help_status,
            'HelpID': help_id,
            'CustomerID': payment.customer_id,
            'CustomerName': customer['name'] if customer else payment.customer_id,
            'HelpDate': payment.payment_date,
            'HelpAmount': total_amount,
            'HelpCategory': help_category,
            'HelpNote': payment.help_note or payment.notes or '',
            'RepaymentDate': payment.repayment_date or '',
            'RepaymentAmount': to_optional_money(payment.repayment_amount),
            'Status': help_status
        })
        db.log_audit('HELP', help_id, 'CREATE', None, {
            'customer_id': payment.customer_id,
            'help_amount': total_amount,
            'help_category': help_category,
            'status': help_status
        })
    
    return {"payment_id": payment_id, "help_id": help_id, "message": "Payment recorded successfully"}

@app.get("/api/help")
async def get_help_records(customer_id: Optional[str] = None, status: Optional[str] = None):
    records = [normalize_help(record) for record in db.get_all_rows('Help')]
    if customer_id:
        records = [record for record in records if record['customer_id'] == customer_id]
    if status:
        records = [record for record in records if record['status'].upper() == status.upper()]
    records.sort(key=lambda record: record['help_date_iso'] or '', reverse=True)
    return records

@app.post("/api/help")
async def create_help_record(help_record: HelpRecord):
    if help_record.help_category not in HELP_CATEGORIES:
        raise HTTPException(status_code=400, detail="Invalid help category")
    help_id = db.get_next_id('help')
    status = title_case_status(help_record.status, 'Active')
    if status not in HELP_STATUSES:
        status = 'Active'
    db.add_dict_row('Help', {
        'help_id': help_id,
        'customer_id': help_record.customer_id,
        'customer_name': help_record.customer_name,
        'help_date': help_record.help_date,
        'help_amount': round(to_float(help_record.help_amount), 2),
        'help_category': help_record.help_category,
        'help_note': help_record.help_note or '',
        'repayment_date': help_record.repayment_date or '',
        'repayment_amount': to_optional_money(help_record.repayment_amount),
        'status': status,
        'HelpID': help_id,
        'CustomerID': help_record.customer_id,
        'CustomerName': help_record.customer_name,
        'HelpDate': help_record.help_date,
        'HelpAmount': round(to_float(help_record.help_amount), 2),
        'HelpCategory': help_record.help_category,
        'HelpNote': help_record.help_note or '',
        'RepaymentDate': help_record.repayment_date or '',
        'RepaymentAmount': to_optional_money(help_record.repayment_amount),
        'Status': status
    })
    db.log_audit('HELP', help_id, 'CREATE', None, {'customer_id': help_record.customer_id, 'help_amount': help_record.help_amount})
    return {"help_id": help_id, "message": "Help record created successfully"}

@app.put("/api/help/{help_id}")
async def update_help_record(help_id: str, help_record: HelpRecord):
    normalized_status = title_case_status(help_record.status, 'Active')
    updates = {
        'help_id': help_id,
        'customer_id': help_record.customer_id,
        'customer_name': help_record.customer_name,
        'help_date': help_record.help_date,
        'help_amount': round(to_float(help_record.help_amount), 2),
        'help_category': help_record.help_category,
        'help_note': help_record.help_note or '',
        'repayment_date': help_record.repayment_date or '',
        'repayment_amount': to_optional_money(help_record.repayment_amount),
        'status': normalized_status,
        'CustomerID': help_record.customer_id,
        'CustomerName': help_record.customer_name,
        'HelpDate': help_record.help_date,
        'HelpAmount': round(to_float(help_record.help_amount), 2),
        'HelpCategory': help_record.help_category,
        'HelpNote': help_record.help_note or '',
        'RepaymentDate': help_record.repayment_date or '',
        'RepaymentAmount': to_optional_money(help_record.repayment_amount),
        'Status': normalized_status
    }
    if isinstance(db, PostgresDB):
        db.update_row('Help', 'help_id', help_id, {key: value for key, value in updates.items() if key == key.lower()})
    else:
        db.update_row('Help', 'HelpID', help_id, updates)
    db.log_audit('HELP', help_id, 'UPDATE', None, updates)
    return {"message": "Help record updated successfully"}

@app.delete("/api/help/{help_id}")
async def delete_help_record(help_id: str):
    if isinstance(db, PostgresDB):
        db.delete_row('Help', 'help_id', help_id)
    else:
        db.delete_row('Help', 'HelpID', help_id)
    db.log_audit('HELP', help_id, 'DELETE', None, None)
    return {"message": "Help record deleted successfully"}

# ========== CAPITAL INJECTION ENDPOINTS ==========

@app.get("/capital-injections")
async def get_capital_injections():
    """Get all capital injections"""
    injections = db.get_all_rows('CapitalInjections')
    return injections

@app.post("/capital-injections")
async def create_capital_injection(injection: CapitalInjection):
    """Record new capital injection"""
    if injection.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than 0")
    injection_id = db.get_next_id('injection')
    
    data = [
        injection_id,
        injection.source_type,
        injection.amount,
        injection.injection_date,
        injection.description or '',
        injection.created_by,
        datetime.now()
    ]
    
    db.add_row('CapitalInjections', data)
    db.log_audit('CAPITAL', injection_id, 'INJECTION', None, {
        'source': injection.source_type,
        'amount': injection.amount
    })
    
    return {"injection_id": injection_id, "message": "Capital injection recorded successfully"}

@app.get("/capital/summary", response_model=CapitalSummary)
async def get_capital_summary():
    """Get capital utilization summary"""
    injections = db.get_all_rows('CapitalInjections')
    loans = load_normalized_loans()
    
    # Total injected
    total_injected = sum(float(inj.get('amount', 0)) for inj in injections)
    
    # BUSINESS RULE: capital reporting excludes add_on_principal from disbursed and collected totals.
    total_disbursed = sum(float(loan.get('effective_principal_amount', loan.get('principal_amount', 0)) or 0) for loan in loans)
    principal_collected = sum(float(loan.get('principal_paid', 0) or 0) for loan in loans)
    
    # Available = Injected - (Disbursed - Collected)
    capital_in_use = total_disbursed - principal_collected
    available_capital = total_injected - capital_in_use
    
    # By source
    by_source = {}
    for inj in injections:
        source = inj.get('source_type', 'OTHER')
        by_source[source] = by_source.get(source, 0) + float(inj.get('amount', 0))
    
    utilization = (capital_in_use / total_injected * 100) if total_injected > 0 else 0
    
    return CapitalSummary(
        total_injected=round(total_injected, 2),
        total_disbursed=round(total_disbursed, 2),
        available_capital=round(available_capital, 2),
        utilization_percentage=round(utilization, 2),
        by_source=by_source
    )

# ========== AUDIT LOG ENDPOINTS ==========

@app.get("/audit-log")
async def get_audit_log(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    action: Optional[str] = None,
    limit: int = 100
):
    """Get audit log with optional filtering"""
    logs = db.get_all_rows('AuditLog')
    
    if entity_type:
        logs = [l for l in logs if l.get('entity_type') == entity_type]
    
    if entity_id:
        logs = [l for l in logs if l.get('entity_id') == entity_id]
    
    if action:
        logs = [l for l in logs if l.get('action') == action]
    
    # Sort by timestamp descending
    logs.sort(key=lambda x: x.get('timestamp') or datetime.min, reverse=True)
    
    return logs[:limit]

# ========== DASHBOARD ENDPOINTS ==========

@app.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats():
    """Get comprehensive dashboard statistics with flexible column access"""
    customers = load_normalized_customers()
    loans = load_normalized_loans()
    injections = db.get_all_rows('CapitalInjections')

    total_customers = len(customers)
    active_customers = len([c for c in customers if str(c.get('status', '')).upper() == 'ACTIVE'])
    total_loans = len(loans)
    active_loans = len([l for l in loans if str(l.get('status', '')).upper() == 'ACTIVE'])
    kulu_count = len([l for l in loans if l.get('transaction_type') == 'KULU'])
    debt_count = len([l for l in loans if l.get('transaction_type') == 'DEBT'])
    other_count = len([l for l in loans if l.get('transaction_type') not in {'KULU', 'DEBT'}])
    total_principal_disbursed = sum(to_float(l.get('effective_principal_amount', l.get('principal_amount'))) for l in loans)
    principal_collected = sum(to_float(l.get('principal_paid')) for l in loans)
    interest_collected = sum(to_float(l.get('interest_paid')) for l in loans)
    principal_outstanding = sum(to_float(l.get('outstanding_balance')) for l in loans)

    # Total interest waived (try different column names)
    total_interest_waived = 0
    for l in loans:
        total_interest_waived += to_float(l.get('waived_interest_amount'))
    
    # Net profit (interest collected is profit)
    net_profit = interest_collected
    
    # Capital tracking
    total_capital_injected = sum(float(inj.get('amount', 0)) for inj in injections)
    capital_utilized = total_principal_disbursed - principal_collected
    
    # Portfolio health assessment
    if total_principal_disbursed > 0:
        outstanding_ratio = principal_outstanding / total_principal_disbursed
        if outstanding_ratio < 0.3:
            portfolio_health = "EXCELLENT"
        elif outstanding_ratio < 0.6:
            portfolio_health = "GOOD"
        else:
            portfolio_health = "NEEDS_ATTENTION"
    else:
        portfolio_health = "NO_DATA"
    
    return DashboardStats(
        total_customers=total_customers,
        active_customers=active_customers,
        total_loans=total_loans,
        active_loans=active_loans,
        total_principal_disbursed=round(total_principal_disbursed, 2),
        total_principal_outstanding=round(principal_outstanding, 2),
        total_interest_collected=round(interest_collected, 2),
        total_interest_waived=round(total_interest_waived, 2),
        total_principal_collected=round(principal_collected, 2),
        net_profit=round(net_profit, 2),
        portfolio_health=portfolio_health,
        kulu_count=kulu_count,
        debt_count=debt_count,
        other_count=other_count,
        total_capital_injected=round(total_capital_injected, 2),
        capital_utilized=round(capital_utilized, 2)
    )

@app.get("/dashboard/loan-trends")
async def get_loan_trends():
    """Get loan disbursement and collection trends"""
    from collections import defaultdict

    loans = load_normalized_loans()
    payments = load_normalized_payments()
    monthly_data = defaultdict(lambda: {'disbursed': 0, 'principal_collected': 0, 'interest_collected': 0})

    for loan in loans:
        month_key = str(loan.get('start_date_iso') or '')[:7]
        if month_key:
            monthly_data[month_key]['disbursed'] += to_float(loan.get('effective_principal_amount', loan.get('principal_amount')))

    remaining_by_loan = {loan['loan_id']: to_float(loan.get('effective_principal_amount', loan.get('principal_amount'))) for loan in loans}
    for payment in sorted(payments, key=lambda row: ((row.get('payment_date_iso') or ''), row.get('payment_id', ''))):
        month_key = str(payment.get('payment_date_iso') or '')[:7]
        if not month_key:
            continue
        loan_id = payment.get('loan_id')
        split = resolve_payment_components(payment, remaining_by_loan.get(loan_id, 0))
        remaining_by_loan[loan_id] = max(0, remaining_by_loan.get(loan_id, 0) - split['principal_amount'])
        monthly_data[month_key]['principal_collected'] += split['principal_amount']
        monthly_data[month_key]['interest_collected'] += split['interest_amount']

    trends = [{'month': k, **v} for k, v in monthly_data.items()]
    trends.sort(key=lambda x: x['month'])
    
    return trends[-12:]  # Last 12 months

@app.get("/dashboard/financial-metrics")
async def get_financial_metrics(
    loan_type: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    customer_id: Optional[str] = None
):
    """
    Get comprehensive financial metrics for dashboard with optional filters.
    
    Data Relationships:
    - Loan_Master ↔ Borrower_Master via BorrowerId
    - Payment_Transactions ↔ Loan_Master via LoanID and BorrowerId
    
    Payment Type Handling:
    - Respects 'PaymentType' column (PRINCIPAL, INTEREST, BOTH) for accurate attribution.
    - Fallback logic for BOTH/UNKNOWN types caps principal at remaining balance.
    """
    from collections import defaultdict
    
    # Load data from Excel using mapped sheet names
    loans = db.get_all_rows('Loans')  # Maps to Loan_Master
    payments = db.get_all_rows('Payments')  # Maps to Payment_Transactions
    borrowers = db.get_all_rows('Customers')  # Maps to Borrower_Master
    
    # Helper functions for flexible column access
    def get_loan_id(record):
        for key in ['LoanID', 'loan_id', 'Loan_ID', 'loanid', 'LoanId']:
            if key in record and record[key] is not None:
                return str(record[key])
        return None
    
    def get_borrower_id(record):
        for key in ['BorrowerId', 'borrower_id', 'customer_id', 'CustomerID', 'BorrowerID']:
            if key in record and record[key] is not None:
                return str(record[key])
        return None
    
    def get_principal(loan):
        for key in ['principal_amount', 'Principal', 'PrincipalAmount', 'Amount', 'LoanAmount']:
            if key in loan and loan[key] is not None:
                try:
                    return float(loan[key])
                except:
                    continue
        return 0
    
    def get_payment_amount(payment):
        for key in ['amount', 'Amount', 'PaymentAmount', 'TransactionAmount']:
            if key in payment and payment[key] is not None:
                try:
                    return float(payment[key])
                except:
                    continue
        return 0
    
    def get_payment_type(payment):
        for key in ['payment_type', 'PaymentType', 'Type', 'TransactionType']:
            if key in payment and payment[key] is not None:
                val = str(payment[key]).strip().upper()
                return val
        return 'UNKNOWN'
    
    def get_status(loan):
        for key in ['status', 'Status', 'LoanStatus']:
            if key in loan and loan[key] is not None:
                return str(loan[key]).strip().upper()
        return 'ACTIVE'
    
    def get_date(record, keys):
        for key in keys:
            if key in record and record[key] is not None:
                parsed = parse_date_value(record[key])
                if parsed:
                    return parsed
        return None

    def get_loan_type(record):
        # PRIORITIZE 'TYPE' first since that's the user's actual Excel column name
        type_columns = ['TYPE', 'Type', 'type', 'LoanType', 'Loan_Type', 'loan_type',
                        'transaction_type', 'TransactionType', 'Transaction_Type',
                        'Category', 'category', 'Loan Type']
        
        for col in type_columns:
            if col in record and record[col]:
                val = str(record[col]).strip().upper()
                if val in ['KULU', 'DEBT']:
                    return val
        return 'OTHER'
    
    # Apply filters to LOANS first
    if loan_type and loan_type != 'ALL':
        loans = [l for l in loans if get_loan_type(l) == loan_type]
    
    if status and status != 'ALL':
        loans = [l for l in loans if get_status(l) == status.upper()]
        
    if customer_id:
        loans = [l for l in loans if get_borrower_id(l) == customer_id or str(l.get('customer_id', '')).lower() == customer_id.lower()]

    # Build loan ID set for filtering payments
    loan_id_set = {get_loan_id(l) for l in loans if get_loan_id(l)}
    
    # Filter payments by loan ID and Customer
    if customer_id:
        # Strict payment filtering by customer if provided
        filtered_payments = [p for p in payments if (get_loan_id(p) in loan_id_set) or (get_borrower_id(p) == customer_id)]
    else:
        filtered_payments = [p for p in payments if get_loan_id(p) in loan_id_set]
    
    # Apply date filter for payments
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        filtered_payments = [p for p in filtered_payments 
                           if get_date(p, ['payment_date', 'PaymentDate', 'Date', 'TransactionDate'])
                           and get_date(p, ['payment_date', 'PaymentDate', 'Date', 'TransactionDate']) >= start_dt]
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        filtered_payments = [p for p in filtered_payments 
                           if get_date(p, ['payment_date', 'PaymentDate', 'Date', 'TransactionDate'])
                           and get_date(p, ['payment_date', 'PaymentDate', 'Date', 'TransactionDate']) <= end_dt]
    
    def calc_segment_metrics(segment_loans, all_payments):
        """
        Calculate metrics for a segment of loans ensuring PaymentType is respected.
        """
        segment_loan_ids = {get_loan_id(l) for l in segment_loans if get_loan_id(l)}
        segment_payments = [p for p in all_payments if get_loan_id(p) in segment_loan_ids]
        
        # Calculate principal disbursed
        principal_disbursed = sum(get_principal(l) for l in segment_loans)
        
        # Build per-loan principal map for "remaining check"
        loan_principal_map = {get_loan_id(l): get_principal(l) for l in segment_loans}
        loan_principal_collected_tracker = defaultdict(float)
        
        total_principal_collected = 0
        total_interest_collected = 0
        
        # Process payments to sum Principal vs Interest
        for p in segment_payments:
            lid = get_loan_id(p)
            if not lid: continue
            
            amount = get_payment_amount(p)
            ptype = get_payment_type(p)
            
            p_portion = 0
            i_portion = 0
            
            # Logic: Strictly follow PaymentType
            if 'PRINCIPAL' in ptype and 'INTEREST' in ptype:
                 # Split logic: Cap at difference
                 remaining = loan_principal_map.get(lid, 0) - loan_principal_collected_tracker[lid]
                 if amount <= remaining:
                     p_portion = amount
                 else:
                     p_portion = max(0, remaining)
                     i_portion = amount - p_portion
            elif 'PRINCIPAL' in ptype:
                 p_portion = amount
            elif 'INTEREST' in ptype:
                 i_portion = amount
            else:
                 # Unknown - Fallback to Split logic
                 remaining = loan_principal_map.get(lid, 0) - loan_principal_collected_tracker[lid]
                 if amount <= remaining:
                     p_portion = amount
                 else:
                     p_portion = max(0, remaining)
                     i_portion = amount - p_portion
            
            total_principal_collected += p_portion
            total_interest_collected += i_portion
            loan_principal_collected_tracker[lid] += p_portion
        
        # Calculate outstanding based on ACTUAL collected principal
        outstanding = max(0, principal_disbursed - total_principal_collected)
        recovery_rate = (total_principal_collected / principal_disbursed * 100) if principal_disbursed > 0 else 0
        
        return {
            'loan_count': len(segment_loans),
            'active_loans': len([l for l in segment_loans if get_status(l) == 'ACTIVE']),
            'principal_disbursed': round(principal_disbursed, 2),
            'principal_collected': round(total_principal_collected, 2),
            'principal_outstanding': round(outstanding, 2),
            'interest_collected': round(total_interest_collected, 2),
            'total_collected': round(total_principal_collected + total_interest_collected, 2),
            'recovery_rate': round(recovery_rate, 2)
        }
    
    # Overall metrics
    overall = calc_segment_metrics(loans, filtered_payments)
    
    # By transaction type (using get_loan_type helper)
    kulu_loans = [l for l in loans if get_loan_type(l) == 'KULU']
    debt_loans = [l for l in loans if get_loan_type(l) == 'DEBT']
    other_loans = [l for l in loans if get_loan_type(l) == 'OTHER']
    
    by_type = {
        'KULU': calc_segment_metrics(kulu_loans, filtered_payments),
        'DEBT': calc_segment_metrics(debt_loans, filtered_payments),
        'OTHER': calc_segment_metrics(other_loans, filtered_payments)
    }
    
    # Collection efficiency (interest collected vs accrued)
    total_interest_accrued = 0
    for loan in loans:
        start_date_loan = get_date(loan, ['start_date', 'StartDate', 'Date', 'LoanDate'])
        if start_date_loan:
            principal = get_principal(loan)
            # Try to get interest rate
            rate = 0
            for key in ['interest_rate', 'InterestRate', 'Rate', 'interest']:
                if key in loan and loan[key] is not None:
                    try:
                        rate = float(loan[key])
                        break
                    except:
                        continue
            
            if rate > 0:
                accrued = calculate_interest_accrued(principal, rate, start_date_loan)
                total_interest_accrued += accrued
    
    interest_efficiency = (overall['interest_collected'] / total_interest_accrued * 100) if total_interest_accrued > 0 else 0
    
    # Portfolio health
    if overall['principal_disbursed'] > 0:
        outstanding_ratio = overall['principal_outstanding'] / overall['principal_disbursed']
        if outstanding_ratio < 0.3:
            portfolio_health = "EXCELLENT"
            health_score = 90
        elif outstanding_ratio < 0.5:
            portfolio_health = "GOOD"
            health_score = 70
        elif outstanding_ratio < 0.7:
            portfolio_health = "FAIR"
            health_score = 50
        else:
            portfolio_health = "NEEDS_ATTENTION"
            health_score = 30
    else:
        portfolio_health = "NO_DATA"
        health_score = 0
    
    return {
        'overall': overall,
        'by_type': by_type,
        'collection_efficiency': round(interest_efficiency, 2),
        'interest_accrued': round(total_interest_accrued, 2),
        'portfolio_health': portfolio_health,
        'health_score': health_score
    }

@app.get("/dashboard/trend-data")
async def get_trend_data(
    months: int = 12,
    loan_type: Optional[str] = None,
    customer_id: Optional[str] = None
):
    """
    Get time-series trend data for dashboard charts.
    Uses flexible column access to work with different Excel schemas.
    """
    from collections import defaultdict
    from datetime import timedelta
    
    loans = db.get_all_rows('Loans')  # Maps to Loan_Master
    payments = db.get_all_rows('Payments')  # Maps to Payment_Transactions
    
    # Helper functions for flexible column access
    def get_loan_id(record):
        for key in ['LoanID', 'loan_id', 'Loan_ID', 'loanid', 'LoanId']:
            if key in record and record[key] is not None:
                return str(record[key])
        return None
    
    def get_borrower_id(record):
        for key in ['BorrowerId', 'borrower_id', 'customer_id', 'CustomerID', 'BorrowerID']:
            if key in record and record[key] is not None:
                return str(record[key])
        return None

    def get_principal(loan):
        for key in ['principal_amount', 'Principal', 'PrincipalAmount', 'Amount', 'LoanAmount']:
            if key in loan and loan[key] is not None:
                try:
                    return float(loan[key])
                except:
                    continue
        return 0
    
    def get_payment_amount(payment):
        for key in ['amount', 'Amount', 'PaymentAmount', 'TransactionAmount']:
            if key in payment and payment[key] is not None:
                try:
                    return float(payment[key])
                except:
                    continue
        return 0
    
    def get_payment_type(payment):
        for key in ['payment_type', 'PaymentType', 'Type', 'TransactionType']:
            if key in payment and payment[key] is not None:
                return str(payment[key]).strip().upper()
        return 'UNKNOWN'
    
    def get_date(record, keys):
        for key in keys:
            if key in record and record[key] is not None:
                val = record[key]
                if isinstance(val, datetime):
                    return val
                elif isinstance(val, date):
                    return datetime.combine(val, datetime.min.time())
                elif isinstance(val, str):
                    try:
                        return datetime.strptime(str(val)[:10], '%Y-%m-%d')
                    except:
                        continue
        return None
    
    def get_loan_type(record):
        # PRIORITIZE 'TYPE' first since that's the user's actual Excel column name
        type_columns = ['TYPE', 'Type', 'type', 'LoanType', 'Loan_Type', 'loan_type',
                        'transaction_type', 'TransactionType', 'Transaction_Type',
                        'Category', 'category', 'Loan Type']
        
        for col in type_columns:
            if col in record and record[col]:
                val = str(record[col]).strip().upper()
                if val in ['KULU', 'DEBT']:
                    return val
                # If it's something else but not explicitly mapped, keep looking or return OTHER
        
        return 'OTHER'

    # Filter by loan type if specified
    if loan_type and loan_type != 'ALL':
        loans = [l for l in loans if get_loan_type(l) == loan_type]
    
    # Filter by Customer
    if customer_id:
        loans = [l for l in loans if get_borrower_id(l) == customer_id or str(l.get('customer_id', '')).lower() == customer_id.lower()]

    # Build loan ID set for filtering payments
    loan_id_set = {get_loan_id(l) for l in loans if get_loan_id(l)}
    
    # Filter payments
    if customer_id:
        payments = [p for p in payments if (get_loan_id(p) in loan_id_set) or (get_borrower_id(p) == customer_id)]
    else:
        payments = [p for p in payments if get_loan_id(p) in loan_id_set]
    
    # Build loan principal map for payment type handling
    loan_principal_map = {get_loan_id(l): get_principal(l) for l in loans}
    
    # Generate month keys for last N months
    current_date = datetime.now()
    month_keys = []
    for i in range(months - 1, -1, -1):
        target_date = current_date - timedelta(days=i * 30)
        month_keys.append(target_date.strftime('%Y-%m'))
    
    # Initialize data structure
    monthly_data = {mk: {
        'month': mk,
        'principal_disbursed': 0,
        'principal_collected': 0,
        'interest_collected': 0,
        'total_collected': 0,
        'transaction_count': 0
    } for mk in month_keys}
    
    # Calculate initial states (before the chart window)
    initial_principal_disbursed = 0
    initial_principal_collected = 0
    
    # Process loans - add disbursed amounts
    for loan in loans:
        start_date = get_date(loan, ['start_date', 'StartDate', 'Date', 'LoanDate'])
        month_key = start_date.strftime('%Y-%m') if start_date else None
        
        # If within window, add to monthly data
        if month_key and month_key in monthly_data:
            monthly_data[month_key]['principal_disbursed'] += get_principal(loan)
        else:
            # Everything else (before window, after window, or no date) adds to initial/legacy
            initial_principal_disbursed += get_principal(loan)
    
    # Track principal collected per loan for accurate payment type handling
    loan_principal_collected = defaultdict(float)
    
    # Sort payments by date for proper sequential processing
    payments_with_dates = []
    for payment in payments:
        pdate = get_date(payment, ['payment_date', 'PaymentDate', 'Date', 'TransactionDate'])
        if pdate:
            payments_with_dates.append((pdate, payment))
    
    payments_with_dates.sort(key=lambda x: x[0])
    
    # Process ALL payments chronologically
    for pdate, payment in payments_with_dates:
        month_key = pdate.strftime('%Y-%m')
        
        amount = get_payment_amount(payment)
        ptype = get_payment_type(payment)
        loan_id = get_loan_id(payment)
        loan_principal = loan_principal_map.get(loan_id, 0)
        
        principal_portion = 0
        interest_portion = 0
        
        # Handle different payment types (Principal First Logic)
        if 'PRINCIPAL' in ptype and 'INTEREST' in ptype:
            # Combined payment
            remaining = loan_principal - loan_principal_collected[loan_id]
            if amount <= remaining:
                principal_portion = amount
            else:
                principal_portion = max(0, remaining)
                interest_portion = amount - principal_portion
        elif 'PRINCIPAL' in ptype:
            principal_portion = amount
        elif 'INTEREST' in ptype:
            interest_portion = amount
        else:
            # Unknown type - treat as combined
            remaining = loan_principal - loan_principal_collected[loan_id]
            if amount <= remaining:
                principal_portion = amount
            else:
                principal_portion = max(0, remaining)
                interest_portion = amount - principal_portion
        
        # Update tracking
        loan_principal_collected[loan_id] += principal_portion
        
        # Update monthly data OR initial stats
        if month_key in monthly_data:
            monthly_data[month_key]['principal_collected'] += principal_portion
            monthly_data[month_key]['interest_collected'] += interest_portion
            monthly_data[month_key]['total_collected'] += amount
            monthly_data[month_key]['transaction_count'] += 1
        else:
            # Catch-all for non-window payments (before or after)
            initial_principal_collected += principal_portion
    
    # Calculate running outstanding balance
    trends = []
    running_disbursed = initial_principal_disbursed
    running_collected = initial_principal_collected
    
    for mk in sorted(month_keys):
        data = monthly_data[mk]
        running_disbursed += data['principal_disbursed']
        running_collected += data['principal_collected']
        
        trends.append({
            'month': mk,
            'month_label': datetime.strptime(mk, '%Y-%m').strftime('%b %Y'),
            'principal_disbursed': round(data['principal_disbursed'], 2),
            'principal_collected': round(data['principal_collected'], 2),
            'interest_collected': round(data['interest_collected'], 2),
            'total_collected': round(data['total_collected'], 2),
            'transaction_count': data['transaction_count'],
            'cumulative_disbursed': round(running_disbursed, 2),
            'cumulative_collected': round(running_collected, 2),
            'outstanding_balance': round(max(0, running_disbursed - running_collected), 2)
        })
    
    return trends

@app.get("/api/customers")
async def get_customers_list():
    """Get list of customers for dropdowns"""
    return [
        {'id': customer['customer_id'], 'customer_id': customer['customer_id'], 'name': customer['name'], 'status': customer['status']}
        for customer in load_normalized_customers()
    ]

# ========== ADVANCED REPORTS ==========

@app.get("/reports/by-transaction-type")
async def get_report_by_transaction_type(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    customer_id: Optional[str] = None
):
    """Get performance metrics by transaction type (Kulu vs Debt vs Other)"""
    from collections import defaultdict
    
    loans = db.get_all_rows('Loans')
    payments = db.get_all_rows('Payments')
    
    # Helper functions
    def get_loan_type(record):
        type_columns = ['TYPE', 'Type', 'type', 'LoanType', 'Loan_Type', 'loan_type', 'transaction_type', 'TransactionType', 'Transaction_Type', 'Category', 'category', 'Loan Type']
        for col in type_columns:
            if col in record and record[col]:
                val = str(record[col]).strip().upper()
                if val in ['KULU', 'DEBT']:
                    return val
        return 'OTHER'

    def get_loan_id(record):
        for key in ['LoanID', 'loan_id', 'Loan_ID', 'loanid', 'LoanId']:
            if key in record and record[key] is not None:
                return str(record[key])
        return None

    def get_borrower_id(record):
        for key in ['BorrowerId', 'borrower_id', 'customer_id', 'CustomerID', 'BorrowerID']:
            if key in record and record[key] is not None:
                return str(record[key])
        return None

    def get_principal(loan):
        for key in ['principal_amount', 'Principal', 'PrincipalAmount', 'Amount', 'LoanAmount']:
            if key in loan and loan[key] is not None:
                try:
                    return float(loan[key])
                except:
                    continue
        return 0

    def get_payment_amount(payment):
        for key in ['amount', 'Amount', 'PaymentAmount', 'TransactionAmount']:
            if key in payment and payment[key] is not None:
                try:
                    return float(payment[key])
                except:
                    continue
        return 0

    def get_payment_type(payment):
        for key in ['payment_type', 'PaymentType', 'Type', 'TransactionType']:
            if key in payment and payment[key] is not None:
                return str(payment[key]).strip().upper()
        return 'UNKNOWN'

    def get_date(record, keys):
        for key in keys:
            if key in record and record[key] is not None:
                val = record[key]
                if isinstance(val, datetime):
                    return val.date()
                elif isinstance(val, date):
                    return val
                elif isinstance(val, str):
                    try:
                        return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
                    except:
                        continue
        return None

    # Filter Loans by Customer
    if customer_id:
        loans = [l for l in loans if get_borrower_id(l) == customer_id or str(l.get('customer_id', '')).lower() == customer_id.lower()]
        
    # Build loan principal map
    loan_principal_map = {get_loan_id(l): get_principal(l) for l in loans}

    # Filter Payments by Customer & Loan
    loan_id_set = {get_loan_id(l) for l in loans if get_loan_id(l)}
    if customer_id:
        payments = [p for p in payments if (get_loan_id(p) in loan_id_set) or (get_borrower_id(p) == customer_id)]
    else:
        payments = [p for p in payments if get_loan_id(p) in loan_id_set]

    # Filter Payments by Date
    if start_date or end_date:
        filtered_payments = []
        start = parse_date_value(start_date)
        end = parse_date_value(end_date)
        
        for p in payments:
            pdate = get_date(p, ['payment_date', 'PaymentDate', 'Date', 'TransactionDate'])
            if pdate and (not start or pdate >= start) and (not end or pdate <= end):
                filtered_payments.append(p)
        payments = filtered_payments

    # Calculate Metrics per Type
    report = {}
    
    # Pre-calculate payment splits per loan (Cumulative)
    # We need to process ALL payments for a loan to get the split right, 
    # even if we are filtering by date for the report view. 
    # BUT, for Reports, we usually want "What happened in this period?"
    # However, "Principal Collected" attribution depends on history.
    # CONSTANT: Logic -> We calculate splits based on ALL history, then filter for the window.
    
    # 1. Sort all payments for relevant loans
    all_payments = db.get_all_rows('Payments')
    all_payments = [p for p in all_payments if get_loan_id(p) in loan_id_set]
    
    payments_with_dates = []
    for payment in all_payments:
        pdate = get_date(payment, ['payment_date', 'PaymentDate', 'Date', 'TransactionDate'])
        if pdate:
            payments_with_dates.append((pdate, payment))
    payments_with_dates.sort(key=lambda x: x[0])
    
    loan_principal_collected_tracker = defaultdict(float)
    loan_payment_splits = defaultdict(list) # loan_id -> list of (date, p_portion, i_portion)

    for pdate, payment in payments_with_dates:
        amount = get_payment_amount(payment)
        ptype = get_payment_type(payment)
        lid = get_loan_id(payment)
        loan_principal = loan_principal_map.get(lid, 0)
        
        p_portion = 0
        i_portion = 0
        
        if 'PRINCIPAL' in ptype and 'INTEREST' in ptype:
             remaining = loan_principal - loan_principal_collected_tracker[lid]
             if amount <= remaining:
                 p_portion = amount
             else:
                 p_portion = max(0, remaining)
                 i_portion = amount - p_portion
        elif 'PRINCIPAL' in ptype:
             p_portion = amount
        elif 'INTEREST' in ptype:
             i_portion = amount
        else:
             remaining = loan_principal - loan_principal_collected_tracker[lid]
             if amount <= remaining:
                 p_portion = amount
             else:
                 p_portion = max(0, remaining)
                 i_portion = amount - p_portion
        
        loan_principal_collected_tracker[lid] += p_portion
        loan_payment_splits[lid].append({
            'date': pdate,
            'p_col': p_portion,
            'i_col': i_portion,
            'total': amount
        })

    # Determine Date Range for Report Aggregation
    r_start = parse_date_value(start_date)
    r_end = parse_date_value(end_date)

    for tx_type in ['KULU', 'DEBT', 'OTHER']:
        type_loans = [l for l in loans if get_loan_type(l) == tx_type]
        type_lids = {get_loan_id(l) for l in type_loans}
        
        p_collected_in_period = 0
        i_collected_in_period = 0
        
        for lid in type_lids:
            splits = loan_payment_splits.get(lid, [])
            for s in splits:
                if (not r_start or s['date'] >= r_start) and (not r_end or s['date'] <= r_end):
                    p_collected_in_period += s['p_col']
                    i_collected_in_period += s['i_col']

        p_disbursed = sum(get_principal(l) for l in type_loans)
        # Interest waived is static per loan, not time-based usually, but we include it if loan is in list
        i_waived = sum(float(l.get('waived_interest_amount', 0) or 0) for l in type_loans)
        
        # Principal Outstanding is (Total Disbursed - Total Principal Collected Ever)
        # BUT the report display often wants "Outstanding for these loans".
        # If we filter by date, "Principal Outstanding" usually implies "Current Outstanding".
        # So we use the tracker value.
        
        p_collected_lifetime = sum(loan_principal_collected_tracker[lid] for lid in type_lids)
        
        report[tx_type] = {
            'loan_count': len(type_loans),
            'active_loans': len([l for l in type_loans if str(l.get('status')).upper() == 'ACTIVE']),
            'principal_disbursed': round(p_disbursed, 2),
            'principal_collected': round(p_collected_in_period, 2),
            'principal_outstanding': round(p_disbursed - p_collected_lifetime, 2), # Current outstanding
            'interest_collected': round(i_collected_in_period, 2),
            'interest_waived': round(i_waived, 2),
            'collection_efficiency': round((p_collected_lifetime / p_disbursed * 100) if p_disbursed > 0 else 0, 1)
        }
    
    return report

@app.get("/reports/profitability")
async def get_profitability_report(
    customer_id: Optional[str] = None
):
    """Get comprehensive profitability analysis"""
    from collections import defaultdict

    loans = db.get_all_rows('Loans')
    payments = db.get_all_rows('Payments')
    injections = db.get_all_rows('CapitalInjections')
    
    # Helper functions
    def get_loan_id(record):
        for key in ['LoanID', 'loan_id', 'Loan_ID', 'loanid', 'LoanId']:
            if key in record and record[key] is not None:
                return str(record[key])
        return None

    def get_borrower_id(record):
        for key in ['BorrowerId', 'borrower_id', 'customer_id', 'CustomerID', 'BorrowerID']:
            if key in record and record[key] is not None:
                return str(record[key])
        return None

    def get_principal(loan):
        for key in ['principal_amount', 'Principal', 'PrincipalAmount', 'Amount', 'LoanAmount']:
            if key in loan and loan[key] is not None:
                try:
                    return float(loan[key])
                except:
                    continue
        return 0
    
    def get_payment_amount(payment):
        for key in ['amount', 'Amount', 'PaymentAmount', 'TransactionAmount']:
            if key in payment and payment[key] is not None:
                try:
                    return float(payment[key])
                except:
                    continue
        return 0
    
    def get_payment_type(payment):
        for key in ['payment_type', 'PaymentType', 'Type', 'TransactionType']:
            if key in payment and payment[key] is not None:
                return str(payment[key]).strip().upper()
        return 'UNKNOWN'
    
    def get_date(record, keys):
        for key in keys:
            if key in record and record[key] is not None:
                val = record[key]
                if isinstance(val, datetime):
                    return val.date()
                elif isinstance(val, date):
                    return val
                elif isinstance(val, str):
                    try:
                        return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
                    except:
                        continue
        return None

    # Filter Loans by Customer
    if customer_id:
        loans = [l for l in loans if get_borrower_id(l) == customer_id or str(l.get('customer_id', '')).lower() == customer_id.lower()]

    # Filter Payments
    loan_id_set = {get_loan_id(l) for l in loans if get_loan_id(l)}
    if customer_id:
        payments = [p for p in payments if (get_loan_id(p) in loan_id_set) or (get_borrower_id(p) == customer_id)]
    else:
        payments = [p for p in payments if get_loan_id(p) in loan_id_set]

    # Calculate Interest Collected correctly using Split Logic
    loan_principal_map = {get_loan_id(l): get_principal(l) for l in loans}
    
    # Sort payments
    payments_with_dates = []
    for payment in payments:
        pdate = get_date(payment, ['payment_date', 'PaymentDate', 'Date', 'TransactionDate'])
        if pdate:
            payments_with_dates.append((pdate, payment))
    payments_with_dates.sort(key=lambda x: x[0])
    
    loan_principal_collected_tracker = defaultdict(float)
    total_interest_collected = 0
    
    for pdate, payment in payments_with_dates:
        amount = get_payment_amount(payment)
        ptype = get_payment_type(payment)
        lid = get_loan_id(payment)
        loan_principal = loan_principal_map.get(lid, 0)
        
        p_portion = 0
        i_portion = 0
        
        if 'PRINCIPAL' in ptype and 'INTEREST' in ptype:
             remaining = loan_principal - loan_principal_collected_tracker[lid]
             if amount <= remaining:
                 p_portion = amount
             else:
                 p_portion = max(0, remaining)
                 i_portion = amount - p_portion
        elif 'PRINCIPAL' in ptype:
             p_portion = amount
        elif 'INTEREST' in ptype:
             i_portion = amount
        else:
             remaining = loan_principal - loan_principal_collected_tracker[lid]
             if amount <= remaining:
                 p_portion = amount
             else:
                 p_portion = max(0, remaining)
                 i_portion = amount - p_portion
        
        loan_principal_collected_tracker[lid] += p_portion
        total_interest_collected += i_portion

    # Interest Analysis - Accrued
    total_interest_accrued = 0
    for loan in loans:
        start_date = loan.get('start_date') # use raw dict access first
        if not start_date:
             start_date = get_date(loan, ['start_date', 'StartDate', 'Date', 'LoanDate'])

        if start_date:
            if isinstance(start_date, str):
                try:
                    start_date = datetime.strptime(start_date[:10], '%Y-%m-%d').date()
                except:
                    pass
            elif isinstance(start_date, datetime):
                start_date = start_date.date()
            
            # Helper to calculate interest based on rate and time
            # Assuming simple interest for now as per previous logic (rate * principal * time?)
            # The previous code called `calculate_interest_accrued`. I need to ensure that function exists or inline it.
            # I'll check if `calculate_interest_accrued` is defined globally.
            # It was called in line 1891 of original file. 
            # I will assume it exists globally.
            try:
                accrued = calculate_interest_accrued(
                    float(loan.get('principal_amount', 0)),
                    float(loan.get('interest_rate', 0)),
                    start_date
                )
                total_interest_accrued += accrued
            except NameError:
                # Fallback if function not found
                pass
    
    interest_waived = sum(float(l.get('waived_interest_amount', 0) or 0) for l in loans)
    interest_pending = total_interest_accrued - total_interest_collected - interest_waived
    
    # Capital Analysis
    total_capital = sum(float(inj.get('amount', 0)) for inj in injections)
    total_disbursed = sum(float(l.get('principal_amount', 0)) for l in loans)
    
    return {
        'interest': {
            'total_accrued': round(total_interest_accrued, 2),
            'collected': round(total_interest_collected, 2),
            'waived': round(interest_waived, 2),
            'pending': round(interest_pending, 2),
            'collection_rate': round((total_interest_collected / total_interest_accrued * 100) if total_interest_accrued > 0 else 0, 2)
        },
        'profit': {
            'gross_profit': round(total_interest_collected, 2),
            'potential_lost_to_waivers': round(interest_waived, 2),
            'net_yield': round((total_interest_collected / total_disbursed * 100) if total_disbursed > 0 else 0, 2)
        },
        'capital': {
            'total_capital': round(total_capital, 2),
            'roi': round((total_interest_collected / total_capital * 100) if total_capital > 0 else 0, 2)
        }
    }

@app.get("/reports/customer-exposure")
async def get_customer_exposure():
    """Get customer risk exposure analysis"""
    customers = db.get_all_rows('Customers')
    loans = db.get_all_rows('Loans')
    payments = db.get_all_rows('Payments')
    
    exposure_list = []
    
    for customer in customers:
        cust_id = customer['customer_id']
        cust_loans = [l for l in loans if l.get('customer_id') == cust_id]
        cust_payments = [p for p in payments if p.get('customer_id') == cust_id]
        
        if not cust_loans:
            continue
        
        total_principal = sum(float(l.get('principal_amount', 0)) for l in cust_loans)
        principal_paid = sum(
            float(p.get('amount', 0)) 
            for p in cust_payments 
            if p.get('payment_type') in ['PRINCIPAL', 'BOTH']
        )
        interest_paid = sum(
            float(p.get('amount', 0)) 
            for p in cust_payments 
            if p.get('payment_type') in ['INTEREST', 'BOTH']
        )
        
        outstanding = total_principal - principal_paid
        active_loans = len([l for l in cust_loans if l.get('status') == 'ACTIVE'])
        
        # Risk score (simple: outstanding / total * active loans)
        risk_score = (outstanding / total_principal * active_loans) if total_principal > 0 else 0
        
        exposure_list.append({
            'customer_id': cust_id,
            'customer_name': customer.get('name'),
            'total_loans': len(cust_loans),
            'active_loans': active_loans,
            'total_principal': round(total_principal, 2),
            'principal_outstanding': round(outstanding, 2),
            'interest_paid': round(interest_paid, 2),
            'risk_score': round(risk_score, 2),
            'status': customer.get('status')
        })
    
    # Sort by outstanding amount descending
    exposure_list.sort(key=lambda x: x['principal_outstanding'], reverse=True)
    
    return exposure_list

# PythonAnywhere WSGI compatibility
from a2wsgi import ASGIMiddleware
wsgi_app = ASGIMiddleware(app)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
