from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Any, Dict, Tuple

import openpyxl

from loan_chain_migration import DEFAULT_DATABASE_URL, LOCAL_SQLITE_PATH, run_migration

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SOURCE_DEFAULT = PROJECT_ROOT / "Diamond Fincorp Master Data.xlsx"
LEGACY_SOURCE_DEFAULT = PROJECT_ROOT / "excel_schema" / "DIAMOND FINANCE DATA.xlsm"
if not SOURCE_DEFAULT.exists():
    SOURCE_DEFAULT = LEGACY_SOURCE_DEFAULT


def to_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def load_workbook_maps(source_file: Path) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    wb = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    loan_sheet = next(name for name in wb.sheetnames if name.strip() == "Loan_Master")
    payment_sheet = "Payment_Transactions"

    loans: Dict[str, Dict[str, Any]] = {}
    for row in wb[loan_sheet].iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        loan_id = str(row[0] or "").strip()
        principal_amount = round(to_float(row[4]), 2)
        add_on_principal = round(max(0.0, to_float(row[10])), 2)
        loans[loan_id] = {
            "loan_id": loan_id,
            "principal_amount": principal_amount,
            "add_on_principal": add_on_principal,
            "fresh_principal": round(max(0.0, principal_amount - add_on_principal), 2),
            "start_date": row[6],
        }

    payments: Dict[str, Dict[str, Any]] = {}
    for row in wb[payment_sheet].iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        payment_id = str(row[0] or "").strip()
        payment_type = str(row[6] or "INTEREST").strip().upper() or "INTEREST"
        amount = round(to_float(row[5]), 2)
        principal_amount = amount if payment_type in {"PRINCIPAL", "BALANCE"} else 0.0
        interest_amount = amount if payment_type == "INTEREST" else 0.0
        if payment_type == "BOTH":
            principal_amount = amount
            interest_amount = 0.0
        payments[payment_id] = {
            "payment_id": payment_id,
            "loan_id": str(row[1] or "").strip(),
            "customer_id": str(row[2] or "").strip(),
            "payment_date": row[4],
            "amount": amount,
            "payment_type": payment_type,
            "principal_amount": principal_amount,
            "interest_amount": interest_amount,
            "is_virtual": 1 if payment_type == "BALANCE" else 0,
        }

    wb.close()
    return loans, payments


def ensure_column(conn: sqlite3.Connection, table_name: str, column_name: str, definition: str) -> None:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    existing = {str(row[1]).lower() for row in rows}
    if column_name.lower() not in existing:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}")


def repair_local_database(source_file: Path) -> Dict[str, Any]:
    workbook_loans, workbook_payments = load_workbook_maps(source_file)
    conn = sqlite3.connect(str(LOCAL_SQLITE_PATH))
    conn.row_factory = sqlite3.Row

    ensure_column(conn, "payment_transactions", "is_virtual", "BOOLEAN DEFAULT 0")
    ensure_column(conn, "payment_transactions", "linked_successor_loan_id", "TEXT")
    ensure_column(conn, "loan_master", "parent_loan_id", "TEXT")
    ensure_column(conn, "loan_master", "loan_chain_id", "TEXT")
    ensure_column(conn, "loan_master", "fresh_principal", "REAL")
    ensure_column(conn, "loan_master", "chain_start_date", "TEXT")

    existing_loan_ids = {str(row[0]) for row in conn.execute("SELECT loan_id FROM loan_master")}
    existing_payment_ids = {str(row[0]) for row in conn.execute("SELECT payment_id FROM payment_transactions")}

    loan_updates = 0
    for loan_id, loan in workbook_loans.items():
        if loan_id not in existing_loan_ids:
            continue
        conn.execute(
            """
            UPDATE loan_master
            SET add_on_principal = ?,
                fresh_principal = ?,
                chain_start_date = COALESCE(chain_start_date, start_date)
            WHERE loan_id = ?
            """,
            (loan["add_on_principal"], loan["fresh_principal"], loan_id),
        )
        loan_updates += 1

    payment_updates = 0
    payment_inserts = 0
    payment_skips = 0
    for payment_id, payment in workbook_payments.items():
        payload = (
            payment["loan_id"],
            payment["customer_id"],
            payment["payment_date"],
            payment["amount"],
            payment["payment_type"],
            payment["principal_amount"],
            payment["interest_amount"],
            payment["is_virtual"],
            payment_id,
        )
        if payment_id in existing_payment_ids:
            conn.execute(
                """
                UPDATE payment_transactions
                SET loan_id = ?,
                    customer_id = ?,
                    payment_date = ?,
                    amount = ?,
                    payment_type = ?,
                    principal_amount = ?,
                    interest_amount = ?,
                    is_virtual = ?
                WHERE payment_id = ?
                """,
                payload,
            )
            payment_updates += 1
            continue

        if payment["loan_id"] not in existing_loan_ids:
            payment_skips += 1
            continue

        conn.execute(
            """
            INSERT INTO payment_transactions (
                payment_id, loan_id, customer_id, payment_date, amount,
                payment_type, payment_method, reference_number, created_date, created_by,
                notes, principal_amount, interest_amount, help_category,
                is_virtual, linked_successor_loan_id
            ) VALUES (?, ?, ?, ?, ?, ?, 'CASH', '', ?, 'SYSTEM', '', ?, ?, 'None', ?, '')
            """,
            (
                payment["payment_id"],
                payment["loan_id"],
                payment["customer_id"],
                payment["payment_date"],
                payment["amount"],
                payment["payment_type"],
                payment["payment_date"],
                payment["principal_amount"],
                payment["interest_amount"],
                payment["is_virtual"],
            ),
        )
        payment_inserts += 1

    conn.execute(
        """
        UPDATE loan_master
        SET fresh_principal = ROUND(
            CASE
                WHEN COALESCE(principal_amount, 0) - COALESCE(add_on_principal, 0) < 0 THEN 0
                ELSE COALESCE(principal_amount, 0) - COALESCE(add_on_principal, 0)
            END,
            2
        ),
            chain_start_date = COALESCE(chain_start_date, start_date)
        """
    )
    conn.commit()

    payment_type_rows = conn.execute(
        """
        SELECT payment_type, COUNT(*) AS row_count, ROUND(COALESCE(SUM(amount), 0), 2) AS total_amount
        FROM payment_transactions
        GROUP BY payment_type
        ORDER BY payment_type
        """
    ).fetchall()
    fresh_total, principal_total = conn.execute(
        "SELECT ROUND(COALESCE(SUM(fresh_principal), 0), 2), ROUND(COALESCE(SUM(principal_amount), 0), 2) FROM loan_master"
    ).fetchone()
    conn.close()

    migration_report = run_migration(DEFAULT_DATABASE_URL)
    return {
        "source_file": str(source_file),
        "loan_rows_updated": loan_updates,
        "payment_rows_updated": payment_updates,
        "payment_rows_inserted": payment_inserts,
        "payment_rows_skipped": payment_skips,
        "payment_type_totals": [tuple(row) for row in payment_type_rows],
        "loan_totals": {
            "fresh_principal_total": fresh_total,
            "principal_amount_total": principal_total,
        },
        "migration_report": migration_report,
    }


if __name__ == "__main__":
    report = repair_local_database(SOURCE_DEFAULT)
    import json
    print(json.dumps(report, indent=2, default=str))
