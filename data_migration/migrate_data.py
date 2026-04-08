from __future__ import annotations

import collections
import contextlib
import hashlib
import io
import json
import os
import re
import shutil
import statistics
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from excel_schema.create_database import create_loan_database

SOURCE_DEFAULT = PROJECT_ROOT / "Diamond Fincorp Master Data.xlsx"
LEGACY_SOURCE_DEFAULT = PROJECT_ROOT / "excel_schema" / "DIAMOND FINANCE DATA.xlsm"
if not SOURCE_DEFAULT.exists():
    SOURCE_DEFAULT = LEGACY_SOURCE_DEFAULT
TARGET_DEFAULT = PROJECT_ROOT / "excel_schema" / "LoanManagement_DB.xlsx"
BACKUP_DIR = PROJECT_ROOT / "excel_schema" / "backups"
LOG_DIR = PROJECT_ROOT / "logs"
VALID_LOAN_STATUSES = {"ACTIVE", "COMPLETED", "HELP", "LOSS", "CLOSED"}
VALID_PAYMENT_TYPES = {"INTEREST", "PRINCIPAL", "BOTH", "BALANCE"}


@dataclass
class MigrationResult:
    summary: Dict[str, Any]
    report_path: Path


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def append_issue(issues: List[Dict[str, Any]], severity: str, code: str, message: str, **details: Any) -> None:
    issues.append(
        {
            "timestamp": now_iso(),
            "severity": severity,
            "code": code,
            "message": message,
            "details": details,
        }
    )


def canonical_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def headers_match(actual_headers: List[Any], expected_columns: List[Iterable[str]]) -> bool:
    for index, aliases in enumerate(expected_columns):
        actual = canonical_header(actual_headers[index] if index < len(actual_headers) else "")
        allowed = {canonical_header(alias) for alias in aliases}
        if actual not in allowed:
            return False
    return True


def to_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def to_datetime(value: Any) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    return None


def normalize_phone(value: Any) -> str:
    phone = str(value or "").strip()
    return re.sub(r"\s+", " ", phone)


def iter_nonempty_rows(ws) -> Tuple[List[Any], List[Tuple[int, Tuple[Any, ...]]], int]:
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows: List[Tuple[int, Tuple[Any, ...]]] = []
    blank_key_rows = 0
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and row[0] not in (None, ""):
            rows.append((row_number, row))
        else:
            blank_key_rows += 1
    return headers, rows, blank_key_rows


def validate_source_schema(wb, issues: List[Dict[str, Any]]) -> None:
    expected = {
        "Borrower_Master ": [
            ("BorrowerID",),
            ("BorrowerName",),
            ("Phone",),
            ("Address",),
            ("IsActive",),
            ("CreatedOn",),
        ],
        "Loan_Master ": [
            ("LoanID",),
            ("BorrowerID",),
            ("Borrower Name",),
            ("TYPE",),
            ("PrincipalAmount",),
            ("InterestRate",),
            ("StartDate",),
            ("FundSourceID",),
            ("LoanStatus",),
            ("CreatedOn",),
            ("ADD ON PRINCIPAL", "AddOnPrincipal"),
            ("REMARKS", "Notes"),
        ],
        "Payment_Transactions": [
            ("PaymentID",),
            ("LoanID",),
            ("BorrowerID", "Borrower"),
            ("BorrowerName", "Borrower Name"),
            ("PaymentDate",),
            ("Amount", "PaymentAmount"),
            ("PaymentType",),
            ("Remarks", "Notes"),
            ("CreatedDate", "CreatedOn"),
            ("Notes", ""),
        ],
    }
    for sheet_name, expected_columns in expected.items():
        if sheet_name not in wb.sheetnames:
            append_issue(issues, "fatal", "MISSING_SHEET", f"Required source sheet '{sheet_name}' is missing.", sheet=sheet_name)
            continue
        ws = wb[sheet_name]
        actual_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        if not headers_match(actual_headers, expected_columns):
            append_issue(
                issues,
                "fatal",
                "SCHEMA_MISMATCH",
                f"Source sheet '{sheet_name}' does not match the expected schema.",
                sheet=sheet_name,
                expected=[list(column) for column in expected_columns],
                actual=actual_headers,
            )


def load_source_records(source_file: Path, issues: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    wb = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    validate_source_schema(wb, issues)

    customers_headers, customer_rows, blank_customers = iter_nonempty_rows(wb["Borrower_Master "])
    loans_headers, loan_rows, blank_loans = iter_nonempty_rows(wb["Loan_Master "])
    payments_headers, payment_rows, blank_payments = iter_nonempty_rows(wb["Payment_Transactions"])

    if blank_loans:
        append_issue(
            issues,
            "warning",
            "USED_RANGE_BLOAT",
            "Loan_Master has a bloated used range. Blank-key rows were ignored during import recovery.",
            sheet="Loan_Master ",
            blank_key_rows=blank_loans,
            max_row=wb["Loan_Master "].max_row,
        )
    if blank_payments:
        append_issue(
            issues,
            "warning",
            "USED_RANGE_BLOAT",
            "Payment_Transactions has a bloated used range. Blank-key rows were ignored during import recovery.",
            sheet="Payment_Transactions",
            blank_key_rows=blank_payments,
            max_row=wb["Payment_Transactions"].max_row,
        )

    customers: List[Dict[str, Any]] = []
    for row_number, row in customer_rows:
        customer_id = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        if not customer_id or not name:
            append_issue(
                issues,
                "warning",
                "CUSTOMER_INCOMPLETE",
                "Customer row is missing a mandatory value and was skipped.",
                row_number=row_number,
                customer_id=customer_id,
            )
            continue
        is_active = str(row[4] or "").strip().lower()
        status = "ACTIVE" if is_active in {"yes", "true", "1"} else "INACTIVE"
        customers.append(
            {
                "customer_id": customer_id,
                "name": name,
                "phone": normalize_phone(row[2]),
                "email": "",
                "address": str(row[3] or "").strip(),
                "id_proof_type": "",
                "id_proof_number": "",
                "status": status,
                "created_date": to_datetime(row[5]) or datetime.now(),
                "notes": "",
                "_source_row": row_number,
            }
        )

    loans_raw: List[Dict[str, Any]] = []
    for row_number, row in loan_rows:
        loan_id = str(row[0] or "").strip()
        customer_id = str(row[1] or "").strip()
        if not loan_id or not customer_id:
            append_issue(
                issues,
                "warning",
                "LOAN_INCOMPLETE",
                "Loan row is missing a mandatory ID and was skipped.",
                row_number=row_number,
                loan_id=loan_id,
                customer_id=customer_id,
            )
            continue
        raw_status = str(row[8] or "ACTIVE").strip().upper() or "ACTIVE"
        if raw_status not in VALID_LOAN_STATUSES:
            append_issue(
                issues,
                "warning",
                "LOAN_STATUS_RECOVERED",
                "Loan status was not recognized and was recovered to ACTIVE.",
                row_number=row_number,
                loan_id=loan_id,
                raw_status=raw_status,
            )
            raw_status = "ACTIVE"
        add_on_principal = to_float(row[10])
        fresh_principal = round(max(0.0, to_float(row[4]) - max(0.0, add_on_principal)), 2)
        loans_raw.append(
            {
                "loan_id": loan_id,
                "customer_id": customer_id,
                "borrower_name": str(row[2] or "").strip(),
                "transaction_type": str(row[3] or "DEBT").strip().upper() or "DEBT",
                "principal_amount": round(to_float(row[4]), 2),
                "add_on_principal": round(max(0.0, add_on_principal), 2),
                "fresh_principal": fresh_principal,
                "interest_rate": round(to_float(row[5]), 6),
                "loan_type": "PERSONAL",
                "start_date": to_datetime(row[6]) or datetime.now(),
                "tenure_months": None,
                "status": raw_status,
                "fund_source": str(row[7] or "").strip(),
                "created_date": to_datetime(row[9]) or to_datetime(row[6]) or datetime.now(),
                "closed_date": None,
                "notes": str(row[11] or "").strip(),
                "debt_interest_mode": "subsequent_collection",
                "pre_deducted_interest": 0.0,
                "net_disbursed_amount": fresh_principal,
                "original_interest_amount": 0.0,
                "waived_interest_amount": 0.0,
                "waiver_reason": "",
                "waiver_date": None,
                "parent_loan_id": "",
                "loan_chain_id": "",
                "chain_start_date": to_datetime(row[6]) or datetime.now(),
                "_source_row": row_number,
            }
        )

    payments_raw: List[Dict[str, Any]] = []
    for row_number, row in payment_rows:
        payment_id = str(row[0] or "").strip()
        loan_id = str(row[1] or "").strip()
        customer_id = str(row[2] or "").strip()
        payment_type = str(row[6] or "INTEREST").strip().upper() or "INTEREST"
        if payment_type not in VALID_PAYMENT_TYPES:
            append_issue(
                issues,
                "warning",
                "PAYMENT_TYPE_RECOVERED",
                "Payment type was not recognized and was recovered to INTEREST.",
                row_number=row_number,
                payment_id=payment_id,
                raw_payment_type=payment_type,
            )
            payment_type = "INTEREST"
        amount = round(to_float(row[5]), 2)
        principal_amount = amount if payment_type in {"PRINCIPAL", "BALANCE"} else 0.0
        interest_amount = amount if payment_type == "INTEREST" else 0.0
        if payment_type == "BOTH":
            principal_amount = amount
            interest_amount = 0.0
        payments_raw.append(
            {
                "payment_id": payment_id,
                "loan_id": loan_id,
                "customer_id": customer_id,
                "payment_date": to_datetime(row[4]) or datetime.now(),
                "amount": amount,
                "payment_type": payment_type,
                "payment_method": "CASH",
                "reference_number": "",
                "created_date": to_datetime(row[8]) or to_datetime(row[4]) or datetime.now(),
                "created_by": "SYSTEM",
                "notes": str((row[9] or row[7] or "")).strip(),
                "principal_amount": principal_amount,
                "interest_amount": interest_amount,
                "help_category": "None",
                "is_virtual": payment_type == "BALANCE",
                "linked_successor_loan_id": "",
                "_source_row": row_number,
            }
        )

    wb.close()
    return {"customers": customers, "loans": loans_raw, "payments": payments_raw}


def dedupe_or_recover_loans(records: Dict[str, List[Dict[str, Any]]], issues: List[Dict[str, Any]]) -> Dict[str, str]:
    loans = records["loans"]
    payments = records["payments"]
    customer_ids = {row["customer_id"] for row in records["customers"]}
    loan_groups: Dict[str, List[Dict[str, Any]]] = collections.defaultdict(list)
    for loan in loans:
        loan_groups[loan["loan_id"]].append(loan)

    numeric_ids = []
    for loan_id in loan_groups.keys():
        match = re.search(r"(\d+)", loan_id)
        if match:
            numeric_ids.append(int(match.group(1)))
    next_numeric = max(numeric_ids or [0]) + 1
    reassigned: Dict[Tuple[str, str], str] = {}
    recovered_id_map: Dict[str, str] = {}

    for loan_id, group in loan_groups.items():
        if len(group) == 1:
            continue
        group.sort(key=lambda item: (item["start_date"], item["customer_id"]))
        customer_group = collections.defaultdict(list)
        for item in group:
            customer_group[item["customer_id"]].append(item)
        if any(len(items) > 1 for items in customer_group.values()):
            append_issue(
                issues,
                "fatal",
                "UNRECOVERABLE_DUPLICATE_LOAN_ID",
                "Duplicate loan IDs could not be recovered because the same customer appears more than once under the same source loan ID.",
                loan_id=loan_id,
                rows=[item["_source_row"] for item in group],
            )
            continue
        for duplicate in group[1:]:
            new_id = f"LN{next_numeric:03d}"
            next_numeric += 1
            reassigned[(loan_id, duplicate["customer_id"])] = new_id
            recovered_id_map[f"{loan_id}:{duplicate['customer_id']}"] = new_id
            append_issue(
                issues,
                "warning",
                "DUPLICATE_LOAN_ID_RECOVERED",
                "Duplicate loan ID was recovered by assigning a new unique loan ID.",
                original_loan_id=loan_id,
                recovered_loan_id=new_id,
                customer_id=duplicate["customer_id"],
                row_number=duplicate["_source_row"],
            )
            duplicate["loan_id"] = new_id

    for payment in payments:
        key = (payment["loan_id"], payment["customer_id"])
        if key in reassigned:
            old_id = payment["loan_id"]
            payment["loan_id"] = reassigned[key]
            append_issue(
                issues,
                "info",
                "PAYMENT_RELINKED",
                "Payment was relinked to a recovered loan ID.",
                payment_id=payment["payment_id"],
                original_loan_id=old_id,
                recovered_loan_id=payment["loan_id"],
                customer_id=payment["customer_id"],
                row_number=payment["_source_row"],
            )

    for loan in loans:
        if loan["customer_id"] not in customer_ids:
            append_issue(
                issues,
                "fatal",
                "LOAN_CUSTOMER_MISSING",
                "Loan references a customer ID that does not exist in the source borrower master.",
                loan_id=loan["loan_id"],
                customer_id=loan["customer_id"],
                row_number=loan["_source_row"],
            )

    loan_ids = {loan["loan_id"] for loan in loans}
    for payment in payments:
        if payment["loan_id"] not in loan_ids:
            append_issue(
                issues,
                "fatal",
                "PAYMENT_LOAN_MISSING",
                "Payment references a loan ID that does not exist after recovery.",
                payment_id=payment["payment_id"],
                loan_id=payment["loan_id"],
                row_number=payment["_source_row"],
            )
        if payment["customer_id"] not in customer_ids:
            append_issue(
                issues,
                "fatal",
                "PAYMENT_CUSTOMER_MISSING",
                "Payment references a customer ID that does not exist in the source borrower master.",
                payment_id=payment["payment_id"],
                customer_id=payment["customer_id"],
                row_number=payment["_source_row"],
            )

    return recovered_id_map


def build_clean_workbook(records: Dict[str, List[Dict[str, Any]]], destination: Path) -> Dict[str, int]:
    with contextlib.redirect_stdout(io.StringIO()):
        create_loan_database(str(destination))
    wb = openpyxl.load_workbook(destination)

    ws_customers = wb["Customers"]
    for row in records["customers"]:
        ws_customers.append(
            [
                row["customer_id"],
                row["name"],
                row["phone"],
                row["email"],
                row["address"],
                row["id_proof_type"],
                row["id_proof_number"],
                row["status"],
                row["created_date"],
                row["notes"],
            ]
        )

    ws_loans = wb["Loans"]
    for row in records["loans"]:
        ws_loans.append(
            [
                row["loan_id"],
                row["customer_id"],
                row["principal_amount"],
                row["add_on_principal"],
                row["fresh_principal"],
                row["interest_rate"],
                row["loan_type"],
                row["start_date"],
                row["tenure_months"],
                row["status"],
                row["fund_source"],
                row["created_date"],
                row["closed_date"],
                row["notes"],
                row["transaction_type"],
                row["debt_interest_mode"],
                row["pre_deducted_interest"],
                row["net_disbursed_amount"],
                row["original_interest_amount"],
                row["waived_interest_amount"],
                row["waiver_reason"],
                row["waiver_date"],
                row["parent_loan_id"],
                row["loan_chain_id"],
                row["chain_start_date"],
            ]
        )

    ws_payments = wb["Payments"]
    for row in records["payments"]:
        ws_payments.append(
            [
                row["payment_id"],
                row["loan_id"],
                row["customer_id"],
                row["payment_date"],
                row["amount"],
                row["payment_type"],
                row["payment_method"],
                row["reference_number"],
                row["created_date"],
                row["created_by"],
                row["notes"],
                row["principal_amount"],
                row["interest_amount"],
                row["help_category"],
                row["is_virtual"],
                row["linked_successor_loan_id"],
            ]
        )

    ws_config = wb["SystemConfig"]
    config_values = {
        "next_customer_id": str(len(records["customers"]) + 1),
        "next_loan_id": str(len(records["loans"]) + 1),
        "next_payment_id": str(len(records["payments"]) + 1),
        "next_help_id": "1",
    }
    for row in ws_config.iter_rows(min_row=2):
        key = str(row[0].value or "")
        if key in config_values:
            row[1].value = config_values[key]
            row[3].value = datetime.now()

    wb.save(destination)
    wb.close()
    return {
        "customers": len(records["customers"]),
        "loans": len(records["loans"]),
        "payments": len(records["payments"]),
    }


def validate_target_integrity(path: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    customers = [row for row in wb["Customers"].iter_rows(min_row=2, values_only=True) if row and row[0] not in (None, "")]
    loans = [row for row in wb["Loans"].iter_rows(min_row=2, values_only=True) if row and row[0] not in (None, "")]
    payments = [row for row in wb["Payments"].iter_rows(min_row=2, values_only=True) if row and row[0] not in (None, "")]
    wb.close()

    customer_ids = {str(row[0]).strip() for row in customers}
    loan_ids = {str(row[0]).strip() for row in loans}
    duplicate_loans = [loan_id for loan_id, count in collections.Counter(str(row[0]).strip() for row in loans).items() if count > 1]
    duplicate_customers = [customer_id for customer_id, count in collections.Counter(str(row[0]).strip() for row in customers).items() if count > 1]
    duplicate_payments = [payment_id for payment_id, count in collections.Counter(str(row[0]).strip() for row in payments).items() if count > 1]

    missing_loan_customers = sorted({str(row[1]).strip() for row in loans if str(row[1]).strip() not in customer_ids})
    missing_payment_loans = sorted({str(row[1]).strip() for row in payments if str(row[1]).strip() not in loan_ids})
    missing_payment_customers = sorted({str(row[2]).strip() for row in payments if str(row[2]).strip() not in customer_ids})

    total_amount_disbursed = round(sum(to_float(row[2]) - max(0.0, to_float(row[3])) for row in loans), 2)

    return {
        "counts": {
            "customers": len(customers),
            "loans": len(loans),
            "payments": len(payments),
        },
        "duplicates": {
            "customers": duplicate_customers,
            "loans": duplicate_loans,
            "payments": duplicate_payments,
        },
        "missing_references": {
            "loan_customers": missing_loan_customers,
            "payment_loans": missing_payment_loans,
            "payment_customers": missing_payment_customers,
        },
        "kpis": {
            "total_amount_disbursed": total_amount_disbursed,
        },
    }


def datetime_key(value: Any) -> str:
    parsed = to_datetime(value)
    if parsed:
        return parsed.isoformat(sep=" ", timespec="seconds")
    return str(value or "").strip()


def build_target_records(path: Path) -> Dict[str, List[Dict[str, Any]]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    customers: List[Dict[str, Any]] = []
    for row in wb["Customers"].iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        customers.append(
            {
                "customer_id": str(row[0] or "").strip(),
                "name": str(row[1] or "").strip(),
                "phone": normalize_phone(row[2]),
                "address": str(row[4] or "").strip(),
                "status": str(row[7] or "").strip().upper(),
                "created_date": row[8],
            }
        )

    loans: List[Dict[str, Any]] = []
    for row in wb["Loans"].iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        loans.append(
            {
                "loan_id": str(row[0] or "").strip(),
                "customer_id": str(row[1] or "").strip(),
                "principal_amount": round(to_float(row[2]), 2),
                "add_on_principal": round(to_float(row[3]), 2),
                "interest_rate": round(to_float(row[4]), 6),
                "start_date": row[6],
                "status": str(row[8] or "").strip().upper(),
                "fund_source": str(row[9] or "").strip(),
                "created_date": row[10],
                "notes": str(row[12] or "").strip(),
                "transaction_type": str(row[13] or "").strip().upper(),
            }
        )

    payments: List[Dict[str, Any]] = []
    for row in wb["Payments"].iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        payments.append(
            {
                "payment_id": str(row[0] or "").strip(),
                "loan_id": str(row[1] or "").strip(),
                "customer_id": str(row[2] or "").strip(),
                "payment_date": row[3],
                "amount": round(to_float(row[4]), 2),
                "payment_type": str(row[5] or "").strip().upper(),
                "created_date": row[8],
                "notes": str(row[10] or "").strip(),
            }
        )

    wb.close()
    return {"customers": customers, "loans": loans, "payments": payments}


def compare_source_to_target(records: Dict[str, List[Dict[str, Any]]], target_file: Path) -> Dict[str, Any]:
    target_records = build_target_records(target_file)
    comparisons = {
        "customers": {
            "key": "customer_id",
            "fields": ["name", "phone", "address", "status", "created_date"],
        },
        "loans": {
            "key": "loan_id",
            "fields": ["customer_id", "principal_amount", "add_on_principal", "interest_rate", "start_date", "status", "fund_source", "created_date", "notes", "transaction_type"],
        },
        "payments": {
            "key": "payment_id",
            "fields": ["loan_id", "customer_id", "payment_date", "amount", "payment_type", "created_date", "notes"],
        },
    }

    summary: Dict[str, Any] = {}
    for entity, config in comparisons.items():
        key_field = config["key"]
        fields = config["fields"]
        source_lookup = {str(item[key_field]).strip(): item for item in records[entity]}
        target_lookup = {str(item[key_field]).strip(): item for item in target_records[entity]}
        missing_in_target = sorted(set(source_lookup) - set(target_lookup))
        extra_in_target = sorted(set(target_lookup) - set(source_lookup))
        field_mismatches: List[Dict[str, Any]] = []

        for record_id in sorted(set(source_lookup).intersection(target_lookup)):
            source_row = source_lookup[record_id]
            target_row = target_lookup[record_id]
            for field in fields:
                source_value = source_row.get(field)
                target_value = target_row.get(field)
                if field.endswith("_date"):
                    source_cmp = datetime_key(source_value)
                    target_cmp = datetime_key(target_value)
                elif isinstance(source_value, (int, float)) or isinstance(target_value, (int, float)):
                    source_cmp = round(to_float(source_value), 6)
                    target_cmp = round(to_float(target_value), 6)
                else:
                    source_cmp = str(source_value or "").strip()
                    target_cmp = str(target_value or "").strip()
                if source_cmp != target_cmp:
                    field_mismatches.append(
                        {
                            "id": record_id,
                            "field": field,
                            "source": source_cmp,
                            "target": target_cmp,
                        }
                    )

        summary[entity] = {
            "source_count": len(source_lookup),
            "target_count": len(target_lookup),
            "missing_in_target": missing_in_target[:20],
            "extra_in_target": extra_in_target[:20],
            "field_mismatch_count": len(field_mismatches),
            "field_mismatches_sample": field_mismatches[:20],
        }
    return summary


def verify_backup_integrity(path: Path) -> Dict[str, Any]:
    validation = validate_target_integrity(path)
    return {
        "path": str(path),
        "sha256": sha256_file(path),
        "counts": validation["counts"],
    }


def create_verified_backup(target_file: Path) -> Dict[str, Any]:
    ensure_dir(BACKUP_DIR)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / f"{target_file.stem}_pre_migration_{timestamp}{target_file.suffix}"
    shutil.copy2(target_file, backup_path)
    return verify_backup_integrity(backup_path)


def run_basic_performance_check(target_file: Path, issues: List[Dict[str, Any]]) -> Dict[str, Any]:
    timings: List[float] = []
    for _ in range(5):
        started = time.perf_counter()
        wb = openpyxl.load_workbook(target_file, data_only=True, read_only=True)
        _ = sum(1 for row in wb["Customers"].iter_rows(min_row=2, values_only=True) if row and row[0] not in (None, ""))
        _ = sum(1 for row in wb["Loans"].iter_rows(min_row=2, values_only=True) if row and row[0] not in (None, ""))
        _ = sum(1 for row in wb["Payments"].iter_rows(min_row=2, values_only=True) if row and row[0] not in (None, ""))
        wb.close()
        timings.append(round(time.perf_counter() - started, 4))

    mean_value = round(statistics.mean(timings), 4)
    p95_value = round(max(timings), 4)
    performance = {
        "runs": timings,
        "mean_seconds": mean_value,
        "p95_seconds": p95_value,
    }
    if p95_value > 1.5:
        append_issue(
            issues,
            "warning",
            "PERFORMANCE_DEGRADED",
            "Basic workbook load stress check exceeded the near-instant target.",
            p95_seconds=p95_value,
            mean_seconds=mean_value,
        )
    return performance


def write_report(summary: Dict[str, Any]) -> Path:
    ensure_dir(LOG_DIR)
    path = LOG_DIR / f"migration_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def migrate_data(source_file: Path = SOURCE_DEFAULT, target_file: Path = TARGET_DEFAULT) -> MigrationResult:
    source_file = Path(source_file)
    target_file = Path(target_file)
    issues: List[Dict[str, Any]] = []
    summary: Dict[str, Any] = {
        "started_at": now_iso(),
        "source_file": str(source_file),
        "target_file": str(target_file),
    }

    if not source_file.exists():
        raise FileNotFoundError(f"Source file not found: {source_file}")
    if not target_file.exists():
        raise FileNotFoundError(f"Target file not found: {target_file}")

    records = load_source_records(source_file, issues)
    recovered_id_map = dedupe_or_recover_loans(records, issues)

    fatal_issues = [issue for issue in issues if issue["severity"] == "fatal"]
    source_counts = {
        "customers": len(records["customers"]),
        "loans": len(records["loans"]),
        "payments": len(records["payments"]),
    }
    summary["source_validation"] = {
        "counts_after_recovery": source_counts,
        "recovered_loan_ids": recovered_id_map,
    }
    if fatal_issues:
        summary["status"] = "aborted"
        summary["issues"] = issues
        summary["finished_at"] = now_iso()
        report_path = write_report(summary)
        return MigrationResult(summary=summary, report_path=report_path)

    backup_info = create_verified_backup(target_file)
    summary["backup"] = backup_info

    with tempfile.TemporaryDirectory(prefix="diamond_migration_") as temp_dir:
        temp_target = Path(temp_dir) / target_file.name
        rebuilt_counts = build_clean_workbook(records, temp_target)
        rebuilt_validation = validate_target_integrity(temp_target)
        if any(rebuilt_validation["duplicates"].values()) or any(rebuilt_validation["missing_references"].values()):
            append_issue(
                issues,
                "fatal",
                "REBUILT_WORKBOOK_INVALID",
                "The rebuilt production workbook still failed integrity checks, so replacement was cancelled.",
                validation=rebuilt_validation,
            )
            summary["status"] = "aborted"
            summary["issues"] = issues
            summary["rebuilt_validation"] = rebuilt_validation
            summary["finished_at"] = now_iso()
            report_path = write_report(summary)
            return MigrationResult(summary=summary, report_path=report_path)

        replacement_copy = target_file.with_suffix(".tmp.xlsx")
        shutil.copy2(temp_target, replacement_copy)
        shutil.move(str(replacement_copy), str(target_file))

    post_validation = validate_target_integrity(target_file)
    exact_comparison = compare_source_to_target(records, target_file)
    if any(
        entity_result["missing_in_target"]
        or entity_result["extra_in_target"]
        or entity_result["field_mismatch_count"]
        for entity_result in exact_comparison.values()
    ):
        append_issue(
            issues,
            "fatal",
            "SOURCE_TARGET_MISMATCH",
            "The rebuilt production workbook does not exactly match the recovered source records.",
            comparison=exact_comparison,
        )
    performance = run_basic_performance_check(target_file, issues)

    summary["status"] = "completed" if not any(issue["severity"] == "fatal" for issue in issues) else "completed_with_findings"
    summary["replacement"] = {
        "mode": "controlled_rebuild_replace",
        "rebuilt_counts": rebuilt_counts,
    }
    summary["post_migration_validation"] = post_validation
    summary["exact_source_target_comparison"] = exact_comparison
    summary["performance"] = performance
    summary["issues"] = issues
    summary["finished_at"] = now_iso()
    report_path = write_report(summary)
    return MigrationResult(summary=summary, report_path=report_path)


def print_console_summary(result: MigrationResult) -> None:
    summary = result.summary
    print("=" * 72)
    print("DIAMOND FINANCE DATA MIGRATION")
    print("=" * 72)
    print(f"Status           : {summary.get('status', 'unknown').upper()}")
    print(f"Source           : {summary['source_file']}")
    print(f"Target           : {summary['target_file']}")
    print(f"Report           : {result.report_path}")
    if "backup" in summary:
        print(f"Backup           : {summary['backup']['path']}")
        print(f"Backup SHA256    : {summary['backup']['sha256']}")
    source_counts = summary.get("source_validation", {}).get("counts_after_recovery", {})
    if source_counts:
        print(
            "Recovered Counts : "
            f"{source_counts.get('customers', 0)} customers, "
            f"{source_counts.get('loans', 0)} loans, "
            f"{source_counts.get('payments', 0)} payments"
        )
    validation = summary.get("post_migration_validation", {})
    if validation:
        counts = validation.get("counts", {})
        print(
            "Production Counts: "
            f"{counts.get('customers', 0)} customers, "
            f"{counts.get('loans', 0)} loans, "
            f"{counts.get('payments', 0)} payments"
        )
        print(f"Total Disbursed  : {validation.get('kpis', {}).get('total_amount_disbursed', 0):,.2f}")
    performance = summary.get("performance", {})
    if performance:
        print(
            "Performance      : "
            f"mean {performance.get('mean_seconds', 0):.4f}s, "
            f"p95 {performance.get('p95_seconds', 0):.4f}s"
        )
    print(f"Issues Logged     : {len(summary.get('issues', []))}")
    print("=" * 72)


if __name__ == "__main__":
    source = Path(os.environ.get("LEGACY_EXCEL_SOURCE_PATH", str(SOURCE_DEFAULT)))
    target = Path(os.environ.get("EXCEL_DB_PATH", str(TARGET_DEFAULT)))
    result = migrate_data(source, target)
    print_console_summary(result)
    if result.summary.get("status") != "completed":
        sys.exit(1)
