"""
Create optimized Excel database structure for loan management system
This serves as the backend database - no formulas, just clean structured data
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def create_loan_database(filename='LoanManagement_DB.xlsx'):
    """Create a clean Excel database with proper schema"""
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # ==================== CUSTOMERS TABLE ====================
    ws_customers = wb.create_sheet('Customers')
    ws_customers.append([
        'customer_id',      # Primary Key (e.g., CUST001)
        'name',             # Customer name
        'phone',            # Contact number
        'email',            # Email address
        'address',          # Physical address
        'id_proof_type',    # ID type (Aadhar, PAN, etc)
        'id_proof_number',  # ID number
        'status',           # ACTIVE/INACTIVE
        'created_date',     # Account creation date
        'notes'             # Additional notes
    ])
    
    # ==================== LOANS TABLE ====================
    ws_loans = wb.create_sheet('Loans')
    ws_loans.append([
        'loan_id',          # Primary Key (e.g., LN001)
        'customer_id',      # Foreign Key to Customers
        'principal_amount', # Original loan amount
        'interest_rate',    # Interest rate (monthly, stored as decimal e.g., 0.02 = 2%)
        'loan_type',        # PERSONAL/BUSINESS/MORTGAGE/OTHER
        'start_date',       # Loan disbursement date
        'tenure_months',    # Expected loan duration
        'status',           # ACTIVE/COMPLETED/DEFAULTED/WRITTEN_OFF
        'fund_source',      # Where money came from
        'created_date',     # Record creation date
        'closed_date',      # Loan closure date
        'notes'             # Additional notes
    ])
    
    # ==================== PAYMENTS TABLE ====================
    ws_payments = wb.create_sheet('Payments')
    ws_payments.append([
        'payment_id',       # Primary Key (e.g., PAY0001)
        'loan_id',          # Foreign Key to Loans
        'customer_id',      # Foreign Key to Customers (denormalized for quick lookup)
        'payment_date',     # Date of payment
        'amount',           # Payment amount
        'payment_type',     # PRINCIPAL/INTEREST/BOTH
        'payment_method',   # CASH/CHEQUE/BANK_TRANSFER/UPI
        'reference_number', # Cheque/transaction reference
        'created_date',     # Record creation timestamp
        'created_by',       # User who entered the payment
        'notes'             # Additional notes
    ])
    
    # ==================== INTEREST_RATE_CHANGES TABLE ====================
    ws_interest = wb.create_sheet('InterestRateChanges')
    ws_interest.append([
        'change_id',        # Primary Key
        'loan_id',          # Foreign Key to Loans
        'old_rate',         # Previous interest rate
        'new_rate',         # New interest rate
        'effective_date',   # Date from which new rate applies
        'reason',           # Reason for change
        'created_date',     # Record creation date
        'created_by'        # User who made the change
    ])
    
    # ==================== LOAN_EVENTS TABLE ====================
    ws_events = wb.create_sheet('LoanEvents')
    ws_events.append([
        'event_id',         # Primary Key
        'loan_id',          # Foreign Key to Loans
        'event_type',       # DISBURSEMENT/RESTRUCTURE/DEFAULT/WAIVER/CLOSURE
        'event_date',       # Date of event
        'amount',           # Amount involved (if applicable)
        'description',      # Event description
        'created_date',     # Record creation date
        'created_by'        # User who logged the event
    ])
    
    # ==================== FUND_SOURCES TABLE ====================
    ws_funds = wb.create_sheet('FundSources')
    ws_funds.append([
        'fund_source_id',   # Primary Key
        'fund_name',        # Source name
        'source_type',      # OWN_CAPITAL/BANK_LOAN/INVESTOR/OTHER
        'total_amount',     # Total amount available
        'interest_cost',    # Cost of capital (if borrowed)
        'status',           # ACTIVE/INACTIVE
        'created_date',     # Record creation date
        'notes'             # Additional notes
    ])
    
    # ==================== SYSTEM_CONFIG TABLE ====================
    ws_config = wb.create_sheet('SystemConfig')
    ws_config.append([
        'config_key',       # Configuration parameter name
        'config_value',     # Configuration value
        'description',      # What this config does
        'last_updated'      # Last update timestamp
    ])
    
    # Add default configurations
    ws_config.append(['default_interest_rate', '0.02', 'Default monthly interest rate (2%)', datetime.now()])
    ws_config.append(['company_name', 'Diamond Fincorp', 'Company name for reports', datetime.now()])
    ws_config.append(['next_customer_id', '1', 'Next customer ID sequence', datetime.now()])
    ws_config.append(['next_loan_id', '1', 'Next loan ID sequence', datetime.now()])
    ws_config.append(['next_payment_id', '1', 'Next payment ID sequence', datetime.now()])
    
    # Format all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(filename)
    print(f"✓ Database created: {filename}")
    print(f"✓ Sheets created: {', '.join(wb.sheetnames)}")
    print(f"\nThis Excel file serves as your database backend.")
    print(f"Do not edit it directly - use the application interface.")
    
    return filename

if __name__ == '__main__':
    create_loan_database('/home/claude/loan_management_system/excel_schema/LoanManagement_DB.xlsx')
